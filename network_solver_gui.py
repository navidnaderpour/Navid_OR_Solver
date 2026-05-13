import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import math
import heapq
from collections import defaultdict, deque

import networkx as nx
import matplotlib.pyplot as plt
from PIL import Image, ImageTk

from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image as RLImage
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import os
from datetime import datetime
from openpyxl.drawing.image import Image as XLImage

# ============================================================
# Shortest Path Solver
# ============================================================

class ShortestPathSolver:

    def __init__(self, edges):
        self.edges = edges
        self.graph = defaultdict(list)
        self.nodes = set()

        for u, v, w in edges:
            self.graph[u].append((v, w))
            self.nodes.add(u)
            self.nodes.add(v)

    # --------------------------------------------------------
    # Dijkstra
    # --------------------------------------------------------

    def dijkstra(self, start, end):

        for u, v, w in self.edges:
            if w < 0:
                raise ValueError(
                    "Dijkstra cannot handle negative weights."
                )

        dist = {node: float('inf') for node in self.nodes}
        prev = {}

        dist[start] = 0

        pq = [(0, start)]

        steps = []

        while pq:

            current_dist, current_node = heapq.heappop(pq)

            if current_dist > dist[current_node]:
                continue

            steps.append(
                f"Visit {current_node} with distance {current_dist}"
            )

            for neighbor, weight in self.graph[current_node]:

                distance = current_dist + weight

                if distance < dist[neighbor]:

                    dist[neighbor] = distance
                    prev[neighbor] = current_node

                    heapq.heappush(
                        pq,
                        (distance, neighbor)
                    )

        if dist[end] == float('inf'):
            raise ValueError("No path exists.")

        path = []

        node = end

        while node != start:
            path.append(node)
            node = prev[node]

        path.append(start)

        path.reverse()

        return {
            "algorithm": "Dijkstra",
            "distance": dist[end],
            "path": path,
            "all_distances": dist,
            "steps": steps
        }

    # --------------------------------------------------------
    # Bellman Ford
    # --------------------------------------------------------

    def bellman_ford(self, start, end):

        dist = {node: float('inf') for node in self.nodes}
        prev = {}

        dist[start] = 0

        steps = []

        for i in range(len(self.nodes) - 1):

            updated = False

            for u, v, w in self.edges:

                if dist[u] != float('inf') and dist[u] + w < dist[v]:

                    dist[v] = dist[u] + w
                    prev[v] = u

                    updated = True

                    steps.append(
                        f"Relax edge {u}->{v}"
                    )

            if not updated:
                break

        # negative cycle check

        for u, v, w in self.edges:

            if dist[u] != float('inf') and dist[u] + w < dist[v]:

                raise ValueError(
                    "Negative cycle detected."
                )

        if dist[end] == float('inf'):
            raise ValueError("No path exists.")

        path = []

        node = end

        while node != start:
            path.append(node)
            node = prev[node]

        path.append(start)

        path.reverse()

        return {
            "algorithm": "Bellman-Ford",
            "distance": dist[end],
            "path": path,
            "all_distances": dist,
            "steps": steps
        }
# ============================================================
# Maximum Flow Solver
# ============================================================

class MaxFlowSolver:

    def __init__(self, edges):

        self.graph = defaultdict(dict)

        for u, v, capacity in edges:

            self.graph[u][v] = capacity

            if v not in self.graph:
                self.graph[v] = {}

    # --------------------------------------------------------
    # BFS
    # --------------------------------------------------------

    def _bfs(self, residual, source, sink, parent):

        visited = set()

        queue = deque([source])

        visited.add(source)

        while queue:

            u = queue.popleft()

            for v in residual[u]:

                if v not in visited and residual[u][v] > 0:

                    visited.add(v)

                    parent[v] = u

                    if v == sink:
                        return True

                    queue.append(v)

        return False

    # --------------------------------------------------------
    # Edmonds Karp
    # --------------------------------------------------------

    def edmonds_karp(self, source, sink):

        residual = defaultdict(dict)

        for u in self.graph:

            for v in self.graph[u]:

                residual[u][v] = self.graph[u][v]

                if u not in residual[v]:
                    residual[v][u] = 0

        parent = {}

        max_flow = 0

        augmenting_paths = []

        while self._bfs(residual, source, sink, parent):

            path_flow = float('inf')

            s = sink

            path = []

            while s != source:

                path.insert(0, s)

                path_flow = min(
                    path_flow,
                    residual[parent[s]][s]
                )

                s = parent[s]

            path.insert(0, source)

            augmenting_paths.append({
                "path": path,
                "flow": path_flow
            })

            max_flow += path_flow

            v = sink

            while v != source:

                u = parent[v]

                residual[u][v] -= path_flow
                residual[v][u] += path_flow

                v = parent[v]

        return {
            "algorithm": "Edmonds-Karp",
            "max_flow": max_flow,
            "paths": augmenting_paths
        }



# ============================================================
# Minimum Spanning Tree Solver
# ============================================================

class MSTSolver:

    def __init__(self, edges):
        self.edges = edges
        self.nodes = set()

        for u, v, w in edges:
            self.nodes.add(u)
            self.nodes.add(v)

    def kruskal(self):

        parent = {}
        rank = {}

        for node in self.nodes:
            parent[node] = node
            rank[node] = 0

        def find(node):
            if parent[node] != node:
                parent[node] = find(parent[node])
            return parent[node]

        def union(a, b):
            root_a = find(a)
            root_b = find(b)

            if root_a == root_b:
                return False

            if rank[root_a] < rank[root_b]:
                parent[root_a] = root_b
            elif rank[root_a] > rank[root_b]:
                parent[root_b] = root_a
            else:
                parent[root_b] = root_a
                rank[root_a] += 1

            return True

        mst_edges = []
        total_weight = 0
        steps = []

        sorted_edges = sorted(self.edges, key=lambda x: x[2])

        for u, v, w in sorted_edges:
            if union(u, v):
                mst_edges.append((u, v, w))
                total_weight += w
                steps.append(f"Add edge {u}-{v} with weight {w}")
            else:
                steps.append(f"Skip edge {u}-{v} to avoid cycle")

        if len(mst_edges) != len(self.nodes) - 1:
            raise ValueError("Graph is disconnected. MST cannot be formed.")

        return {
            "algorithm": "Kruskal",
            "total_weight": total_weight,
            "mst_edges": mst_edges,
            "steps": steps
        }

    def prim(self):

        graph = defaultdict(list)

        for u, v, w in self.edges:
            graph[u].append((w, u, v))
            graph[v].append((w, v, u))

        start = next(iter(self.nodes))

        visited = set([start])
        pq = []

        for edge in graph[start]:
            heapq.heappush(pq, edge)

        mst_edges = []
        total_weight = 0
        steps = []

        while pq and len(visited) < len(self.nodes):

            w, u, v = heapq.heappop(pq)

            if v in visited:
                steps.append(f"Skip edge {u}-{v} to avoid cycle")
                continue

            visited.add(v)
            mst_edges.append((u, v, w))
            total_weight += w

            steps.append(f"Add edge {u}-{v} with weight {w}")

            for next_edge in graph[v]:
                if next_edge[2] not in visited:
                    heapq.heappush(pq, next_edge)

        if len(mst_edges) != len(self.nodes) - 1:
            raise ValueError("Graph is disconnected. MST cannot be formed.")

        return {
            "algorithm": "Prim",
            "total_weight": total_weight,
            "mst_edges": mst_edges,
            "steps": steps
        }

# ============================================================
# Network GUI
# ============================================================

class NetworkGUI:

    BG_COLOR = "#f4f6f9"
    HEADER_COLOR = "#0d47a1"
    BUTTON_COLOR = "#1565c0"

    CARD_COLOR = "#ffffff"

    SUCCESS_COLOR = "#2e7d32"

    TABLE_HEADER = "#e3f2fd"

    def __init__(self, root):

        self.root = root

        self.root.title("Network Optimization Solver")

        self.root.geometry("1350x900")

        self.root.configure(bg=self.BG_COLOR)

        self.problem_type = tk.StringVar(value="shortest_path")

        self._setup_styles()

        self._create_header()

        self._create_notebook()

        self._create_page1()

    # --------------------------------------------------------
    # Styles
    # --------------------------------------------------------

    def _setup_styles(self):

        style = ttk.Style()

        style.theme_use("clam")

        style.configure(
            "TNotebook",
            background=self.BG_COLOR,
            borderwidth=0
        )

        style.configure(
            "TNotebook.Tab",
            padding=[18, 8],
            font=("Segoe UI", 10, "bold")
        )

        style.map(
            "TNotebook.Tab",
            background=[("selected", "#1565c0")],
            foreground=[("selected", "white")]
        )

    # --------------------------------------------------------
    # Header
    # --------------------------------------------------------

    def _create_header(self):

        header = tk.Frame(
            self.root,
            bg=self.HEADER_COLOR,
            height=70
        )

        header.pack(fill=tk.X)

        header.pack_propagate(False)

        tk.Label(
            header,
            text="Network Optimization Solver",
            font=("Segoe UI", 20, "bold"),
            fg="white",
            bg=self.HEADER_COLOR
        ).pack(side=tk.LEFT, padx=20)

        tk.Label(
            header,
            text="Prepared by Navid Naderpour",
            font=("Segoe UI", 10),
            fg="#bbdefb",
            bg=self.HEADER_COLOR
        ).pack(side=tk.RIGHT, padx=20)

    # --------------------------------------------------------
    # Notebook
    # --------------------------------------------------------

    def _create_notebook(self):

        self.notebook = ttk.Notebook(self.root)

        self.notebook.pack(
            fill=tk.BOTH,
            expand=True,
            padx=10,
            pady=10
        )

        self.page1 = tk.Frame(self.notebook, bg=self.BG_COLOR)
        self.page2 = tk.Frame(self.notebook, bg=self.BG_COLOR)
        self.page3 = tk.Frame(self.notebook, bg=self.BG_COLOR)
        self.page4 = tk.Frame(self.notebook, bg=self.BG_COLOR)
        self.page5 = tk.Frame(self.notebook, bg=self.BG_COLOR)

        self.notebook.add(self.page1, text=" Setup ")
        self.notebook.add(self.page2, text=" Data Entry ")
        self.notebook.add(self.page3, text=" Solve ")
        self.notebook.add(self.page4, text=" Results ")
        self.notebook.add(self.page5, text=" Visualization ")
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
    
    def _on_tab_changed(self, event):

        selected_tab = self.notebook.select()
        selected_widget = self.root.nametowidget(selected_tab)

        if selected_widget == self.page2:
            if not self.page2.winfo_children():
                self._create_page2()

        elif selected_widget == self.page5:
            if hasattr(self, "result"):
                self._create_page5()
            else:
                for widget in self.page5.winfo_children():
                    widget.destroy()

                frame = tk.Frame(self.page5, bg=self.BG_COLOR)
                frame.pack(expand=True)

                tk.Label(
                    frame,
                    text="Please solve a network problem first.",
                    font=("Segoe UI", 18, "bold"),
                    bg=self.BG_COLOR,
                    fg=self.HEADER_COLOR
                ).pack(pady=20)

                tk.Button(
                    frame,
                    text="Go to Setup",
                    font=("Segoe UI", 12, "bold"),
                    bg=self.BUTTON_COLOR,
                    fg="white",
                    padx=25,
                    pady=8,
                    command=lambda: self.notebook.select(self.page1)
                ).pack()


    # --------------------------------------------------------
    # Page 1
    # --------------------------------------------------------

    def _create_page1(self):

        frame = tk.Frame(
            self.page1,
            bg=self.BG_COLOR
        )

        frame.pack(
            expand=True,
            pady=80
        )

        tk.Label(
            frame,
            text="Select Network Problem Type",
            font=("Segoe UI", 22, "bold"),
            bg=self.BG_COLOR,
            fg=self.HEADER_COLOR
        ).pack(pady=(0, 40))

        card = tk.Frame(
            frame,
            bg=self.CARD_COLOR,
            bd=1,
            relief=tk.SOLID
        )

        card.pack(
            ipadx=40,
            ipady=30
        )

        problems = [

            ("Shortest Path", "shortest_path"),

            ("Maximum Flow", "max_flow"),

            ("Minimum Spanning Tree", "mst"),

            ("Transportation Network", "transport")
        ]

        for text, value in problems:

            row = tk.Frame(
                card,
                bg=self.CARD_COLOR
            )

            row.pack(
                fill=tk.X,
                pady=10
            )

            rb = tk.Radiobutton(
                row,
                text=text,
                variable=self.problem_type,
                value=value,
                font=("Segoe UI", 13),
                bg=self.CARD_COLOR,
                anchor="w",
                width=30
            )

            rb.pack(side=tk.LEFT)

        tk.Button(
            frame,
            text="NEXT →",
            font=("Segoe UI", 14, "bold"),
            bg=self.BUTTON_COLOR,
            fg="white",
            padx=40,
            pady=10,
            command=self._go_to_page2
        ).pack(pady=40)

    # --------------------------------------------------------
    # Go Page 2
    # --------------------------------------------------------

    def _go_to_page2(self):

        self._create_page2()

        self.notebook.select(self.page2)

    # --------------------------------------------------------
    # Page 2
    # --------------------------------------------------------

    def _create_page2(self):

        for widget in self.page2.winfo_children():
            widget.destroy()

        frame = tk.Frame(self.page2, bg=self.BG_COLOR)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        tk.Label(
            frame,
            text="Data Entry",
            font=("Segoe UI", 20, "bold"),
            bg=self.BG_COLOR,
            fg=self.HEADER_COLOR
        ).pack(anchor="w")

        problem_names = {
            "shortest_path": "Shortest Path",
            "max_flow": "Maximum Flow",
            "mst": "Minimum Spanning Tree",
            "transport": "Transportation Network"
        }

        info = tk.Frame(frame, bg="#e3f2fd", bd=1, relief=tk.SOLID)
        info.pack(fill=tk.X, pady=15, ipady=8)

        tk.Label(
            info,
            text="Selected Problem: " + problem_names[self.problem_type.get()],
            font=("Segoe UI", 12, "bold"),
            bg="#e3f2fd",
            fg="#1565c0"
        ).pack()

        form = tk.Frame(frame, bg=self.CARD_COLOR, bd=1, relief=tk.SOLID)
        form.pack(fill=tk.X, pady=10, ipadx=15, ipady=15)

        top_row = tk.Frame(form, bg=self.CARD_COLOR)
        top_row.pack(fill=tk.X, pady=8)

        if self.problem_type.get() == "shortest_path":
            start_label = "Start Node:"
            end_label = "End Node:"
            edge_label = "Edges: From,To,Cost"
            sample_text = (
                "A,B,4\n"
                "A,C,2\n"
                "C,B,1\n"
                "B,D,5\n"
                "C,D,8"
            )
            default_start = "A"
            default_end = "D"

        elif self.problem_type.get() == "max_flow":
            start_label = "Source:"
            end_label = "Sink:"
            edge_label = "Edges: From,To,Capacity"
            sample_text = (
                "S,A,10\n"
                "S,B,5\n"
                "A,B,15\n"
                "A,T,10\n"
                "B,T,10"
            )
            default_start = "S"
            default_end = "T"

        elif self.problem_type.get() == "mst":
            start_label = "Start Node:"
            end_label = "End Node:"
            edge_label = "Edges: Node1,Node2,Weight"
            sample_text = (
                "A,B,4\n"
                "A,C,2\n"
                "B,C,1\n"
                "B,D,5\n"
                "C,D,8\n"
                "C,E,10\n"
                "D,E,2\n"
                "B,E,9"
            )
            default_start = "-"
            default_end = "-"

        else:
            start_label = "Start:"
            end_label = "End:"
            edge_label = "Edges:"
            sample_text = ""
            default_start = ""
            default_end = ""

        tk.Label(
            top_row,
            text=start_label,
            font=("Segoe UI", 11, "bold"),
            bg=self.CARD_COLOR
        ).pack(side=tk.LEFT, padx=(10, 5))

        self.start_entry = tk.Entry(
            top_row,
            width=15,
            font=("Segoe UI", 11),
            justify=tk.CENTER
        )
        self.start_entry.pack(side=tk.LEFT, padx=5)
        self.start_entry.insert(0, default_start)

        tk.Label(
            top_row,
            text=end_label,
            font=("Segoe UI", 11, "bold"),
            bg=self.CARD_COLOR
        ).pack(side=tk.LEFT, padx=(25, 5))

        self.end_entry = tk.Entry(
            top_row,
            width=15,
            font=("Segoe UI", 11),
            justify=tk.CENTER
        )
        self.end_entry.pack(side=tk.LEFT, padx=5)
        self.end_entry.insert(0, default_end)

        if self.problem_type.get() == "shortest_path":
            self.directed_var = tk.BooleanVar(value=True)

            tk.Checkbutton(
                top_row,
                text="Directed Graph",
                variable=self.directed_var,
                font=("Segoe UI", 10),
                bg=self.CARD_COLOR
            ).pack(side=tk.LEFT, padx=30)
        else:
            self.directed_var = tk.BooleanVar(value=True)

        edge_frame = tk.Frame(form, bg=self.CARD_COLOR)
        edge_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        tk.Label(
            edge_frame,
            text=edge_label,
            font=("Segoe UI", 12, "bold"),
            bg=self.CARD_COLOR,
            fg=self.HEADER_COLOR
        ).pack(anchor="w", padx=10, pady=(5, 8))

        self.edge_text = tk.Text(
            edge_frame,
            height=12,
            font=("Consolas", 11),
            bd=1,
            relief=tk.SOLID
        )
        self.edge_text.pack(fill=tk.BOTH, expand=True, padx=10)

        self.edge_text.insert("1.0", sample_text)

        hint_text = (
            "Enter one edge per line. Example: A,B,4\n"
            "Node names can be letters or words. Costs/capacities must be numbers."
        )

        tk.Label(
            edge_frame,
            text=hint_text,
            font=("Segoe UI", 9),
            bg=self.CARD_COLOR,
            fg="#757575",
            justify=tk.LEFT
        ).pack(anchor="w", padx=10, pady=8)

        btn_frame = tk.Frame(frame, bg=self.BG_COLOR)
        btn_frame.pack(pady=20)

        tk.Button(
            btn_frame,
            text="← Back",
            font=("Segoe UI", 11),
            bg="#757575",
            fg="white",
            padx=20,
            pady=7,
            command=lambda: self.notebook.select(self.page1)
        ).pack(side=tk.LEFT, padx=8)

        tk.Button(
            btn_frame,
            text="Load Example",
            font=("Segoe UI", 11, "bold"),
            bg="#00838f",
            fg="white",
            padx=20,
            pady=7,
            command=self._load_example
        ).pack(side=tk.LEFT, padx=8)

        tk.Button(
            btn_frame,
            text="Clear",
            font=("Segoe UI", 11, "bold"),
            bg="#c62828",
            fg="white",
            padx=20,
            pady=7,
            command=self._clear_data_entry
        ).pack(side=tk.LEFT, padx=8)

        tk.Button(
            btn_frame,
            text="Validate & Continue →",
            font=("Segoe UI", 12, "bold"),
            bg=self.BUTTON_COLOR,
            fg="white",
            padx=28,
            pady=8,
            command=self._validate_data_and_go
        ).pack(side=tk.LEFT, padx=8)
    def _load_example(self):

        self.edge_text.delete("1.0", tk.END)

        if self.problem_type.get() == "shortest_path":
            self.start_entry.delete(0, tk.END)
            self.end_entry.delete(0, tk.END)
            self.start_entry.insert(0, "A")
            self.end_entry.insert(0, "D")

            self.edge_text.insert(
                "1.0",
                "A,B,4\n"
                "A,C,2\n"
                "C,B,1\n"
                "B,D,5\n"
                "C,D,8"
            )

        elif self.problem_type.get() == "max_flow":
            self.start_entry.delete(0, tk.END)
            self.end_entry.delete(0, tk.END)
            self.start_entry.insert(0, "S")
            self.end_entry.insert(0, "T")

            self.edge_text.insert(
                "1.0",
                "S,A,10\n"
                "S,B,5\n"
                "A,B,15\n"
                "A,T,10\n"
                "B,T,10"
            )
        elif self.problem_type.get() == "mst":
            self.start_entry.delete(0, tk.END)
            self.end_entry.delete(0, tk.END)
            self.start_entry.insert(0, "-")
            self.end_entry.insert(0, "-")

            self.edge_text.insert(
                "1.0",
                "A,B,4\n"
                "A,C,2\n"
                "B,C,1\n"
                "B,D,5\n"
                "C,D,8\n"
                "C,E,10\n"
                "D,E,2\n"
                "B,E,9"
            )
    def _clear_data_entry(self):

        self.start_entry.delete(0, tk.END)
        self.end_entry.delete(0, tk.END)
        self.edge_text.delete("1.0", tk.END)

    def _parse_edges(self):

        raw_text = self.edge_text.get("1.0", tk.END).strip()

        if not raw_text:
            raise ValueError("Edge list is empty.")

        edges = []
        line_number = 0

        for line in raw_text.splitlines():

            line_number += 1
            line = line.strip()

            if not line:
                continue

            parts = [p.strip() for p in line.split(",")]

            if len(parts) != 3:
                raise ValueError(
                    f"Line {line_number} is invalid. Use format: From,To,Value"
                )

            u, v, value = parts

            if not u or not v:
                raise ValueError(
                    f"Line {line_number}: node names cannot be empty."
                )

            try:
                value = float(value)
            except:
                raise ValueError(
                    f"Line {line_number}: cost/capacity must be numeric."
                )

            if self.problem_type.get() == "max_flow" and value < 0:
                raise ValueError(
                    f"Line {line_number}: capacity cannot be negative."
                )

            if self.problem_type.get() == "shortest_path":
                edges.append((u, v, value))

                if not self.directed_var.get():
                    edges.append((v, u, value))
            else:
                edges.append((u, v, value))

        if not edges:
            raise ValueError("No valid edges found.")

        return edges

    def _validate_data_and_go(self):

        try:
            self.start_node = self.start_entry.get().strip()
            self.end_node = self.end_entry.get().strip()

            if self.problem_type.get() != "mst":

                if not self.start_node:
                    raise ValueError("Start/Source node is empty.")

                if not self.end_node:
                    raise ValueError("End/Sink node is empty.")

                if self.start_node == self.end_node:
                    raise ValueError("Start and End nodes cannot be the same.")

            self.edges = self._parse_edges()

            all_nodes = set()
            for u, v, value in self.edges:
                all_nodes.add(u)
                all_nodes.add(v)

            if self.problem_type.get() != "mst":

                if self.start_node not in all_nodes:
                    raise ValueError("Start/Source node does not exist in edge list.")

                if self.end_node not in all_nodes:
                    raise ValueError("End/Sink node does not exist in edge list.")

            if self.problem_type.get() == "shortest_path":
                has_negative = any(w < 0 for _, _, w in self.edges)
                self.has_negative_weight = has_negative
            else:
                self.has_negative_weight = False

            self._create_page3()
            self.notebook.select(self.page3)

        except Exception as e:
            messagebox.showerror("Input Error", str(e))

    def _create_page3(self):

        for widget in self.page3.winfo_children():
            widget.destroy()

        frame = tk.Frame(
            self.page3,
            bg=self.BG_COLOR
        )

        frame.pack(
            fill=tk.BOTH,
            expand=True,
            padx=20,
            pady=20
        )

        tk.Label(
            frame,
            text="Solve Network Problem",
            font=("Segoe UI", 22, "bold"),
            bg=self.BG_COLOR,
            fg=self.HEADER_COLOR
        ).pack(anchor="w", pady=(0, 20))

        summary = tk.Frame(
            frame,
            bg="#e3f2fd",
            bd=1,
            relief=tk.SOLID
        )

        summary.pack(
            fill=tk.X,
            pady=10,
            ipady=10
        )

        

        if self.problem_type.get() == "mst":
            summary_text = "Problem: Minimum Spanning Tree"
        else:
            summary_text = f"Start: {self.start_node}    |    End: {self.end_node}"

        tk.Label(
            summary,
            text=summary_text,
            font=("Segoe UI", 12, "bold"),
            bg="#e3f2fd",
            fg="#1565c0"
        ).pack()
        tk.Label(
            summary,
            text=f"Edges Entered: {len(self.edges)}",
            font=("Segoe UI", 11),
            bg="#e3f2fd",
            fg="#37474f"
        ).pack()
        # ----------------------------------------------------
        # Algorithm Selection
        # ----------------------------------------------------

        card = tk.Frame(
            frame,
            bg=self.CARD_COLOR,
            bd=1,
            relief=tk.SOLID
        )

        card.pack(
            pady=30,
            ipadx=30,
            ipady=25
        )

        tk.Label(
            card,
            text="Select Algorithm",
            font=("Segoe UI", 18, "bold"),
            bg=self.CARD_COLOR,
            fg=self.HEADER_COLOR
        ).pack(pady=(0, 20))

        self.algorithm_var = tk.StringVar()

        if self.problem_type.get() == "shortest_path":

            self.algorithm_var.set("dijkstra")

            tk.Radiobutton(
                card,
                text="Dijkstra Algorithm",
                variable=self.algorithm_var,
                value="dijkstra",
                font=("Segoe UI", 12),
                bg=self.CARD_COLOR
            ).pack(anchor="w", pady=5)

            tk.Radiobutton(
                card,
                text="Bellman-Ford Algorithm",
                variable=self.algorithm_var,
                value="bellman_ford",
                font=("Segoe UI", 12),
                bg=self.CARD_COLOR
            ).pack(anchor="w", pady=5)

            if self.has_negative_weight:

                warning = tk.Label(
                    card,
                    text="WARNING: Negative edges detected. Use Bellman-Ford.",
                    font=("Segoe UI", 10, "bold"),
                    bg=self.CARD_COLOR,
                    fg="#c62828"
                )

                warning.pack(pady=10)

        elif self.problem_type.get() == "max_flow":

            self.algorithm_var.set("edmonds_karp")

            tk.Radiobutton(
                card,
                text="Edmonds-Karp Algorithm",
                variable=self.algorithm_var,
                value="edmonds_karp",
                font=("Segoe UI", 12),
                bg=self.CARD_COLOR
            ).pack(anchor="w", pady=5)
        elif self.problem_type.get() == "mst":

            self.algorithm_var.set("kruskal")

            tk.Radiobutton(
                card,
                text="Kruskal Algorithm",
                variable=self.algorithm_var,
                value="kruskal",
                font=("Segoe UI", 12),
                bg=self.CARD_COLOR
            ).pack(anchor="w", pady=5)

            tk.Radiobutton(
                card,
                text="Prim Algorithm",
                variable=self.algorithm_var,
                value="prim",
                font=("Segoe UI", 12),
                bg=self.CARD_COLOR
            ).pack(anchor="w", pady=5)
        # ----------------------------------------------------
        # Buttons
        # ----------------------------------------------------

        btn_frame = tk.Frame(
            frame,
            bg=self.BG_COLOR
        )

        btn_frame.pack(pady=30)

        tk.Button(
            btn_frame,
            text="← Back",
            font=("Segoe UI", 11),
            bg="#757575",
            fg="white",
            padx=20,
            pady=7,
            command=lambda: self.notebook.select(self.page2)
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="SOLVE",
            font=("Segoe UI", 15, "bold"),
            bg="#2e7d32",
            fg="white",
            padx=45,
            pady=12,
            command=self._solve_problem
        ).pack(side=tk.LEFT, padx=10)   
    def _solve_problem(self):

        try:

            algorithm = self.algorithm_var.get()

            # ------------------------------------------------
            # Shortest Path
            # ------------------------------------------------

            if self.problem_type.get() == "shortest_path":

                solver = ShortestPathSolver(self.edges)

                if algorithm == "dijkstra":

                    result = solver.dijkstra(
                        self.start_node,
                        self.end_node
                    )

                elif algorithm == "bellman_ford":

                    result = solver.bellman_ford(
                        self.start_node,
                        self.end_node
                    )

                else:
                    raise ValueError("Unknown algorithm.")

            # ------------------------------------------------
            # Maximum Flow
            # ------------------------------------------------

            elif self.problem_type.get() == "max_flow":

                solver = MaxFlowSolver(self.edges)

                if algorithm == "edmonds_karp":

                    result = solver.edmonds_karp(
                        self.start_node,
                        self.end_node
                    )

                else:
                    raise ValueError("Unknown algorithm.")
            elif self.problem_type.get() == "mst":

                solver = MSTSolver(self.edges)

                if algorithm == "kruskal":
                    result = solver.kruskal()

                elif algorithm == "prim":
                    result = solver.prim()

                else:
                    raise ValueError("Unknown algorithm.")
            else:

                raise ValueError(
                    "This problem type is not implemented yet."
                )

            self.result = result

            self._create_page4()

            self.notebook.select(self.page4)

        except Exception as e:

            messagebox.showerror(
                "Solver Error",
                str(e)
            )

        
    # --------------------------------------------------------
    # Page 4 - Results
    # --------------------------------------------------------

    def _create_page4(self):

        for widget in self.page4.winfo_children():
            widget.destroy()

        main = tk.Frame(
            self.page4,
            bg=self.BG_COLOR
        )

        main.pack(
            fill=tk.BOTH,
            expand=True,
            padx=20,
            pady=20
        )

        # ----------------------------------------------------
        # Title
        # ----------------------------------------------------

        tk.Label(
            main,
            text="Optimization Results",
            font=("Segoe UI", 22, "bold"),
            bg=self.BG_COLOR,
            fg=self.HEADER_COLOR
        ).pack(anchor="w", pady=(0, 20))

        # ----------------------------------------------------
        # Summary Card
        # ----------------------------------------------------

        card = tk.Frame(
            main,
            bg="#e8f5e9",
            bd=1,
            relief=tk.SOLID
        )

        card.pack(
            fill=tk.X,
            pady=10,
            ipady=15
        )

        if self.problem_type.get() == "shortest_path":

            algo = self.result["algorithm"]

            distance = self.result["distance"]

            path_text = " → ".join(self.result["path"])

            tk.Label(
                card,
                text=f"{algo} Solution",
                font=("Segoe UI", 18, "bold"),
                bg="#e8f5e9",
                fg="#1b5e20"
            ).pack(pady=5)

            tk.Label(
                card,
                text=f"Shortest Distance = {distance}",
                font=("Segoe UI", 15, "bold"),
                bg="#e8f5e9",
                fg="#2e7d32"
            ).pack(pady=5)

            tk.Label(
                card,
                text=f"Path: {path_text}",
                font=("Segoe UI", 13),
                bg="#e8f5e9",
                fg="#000000"
            ).pack(pady=5)

        elif self.problem_type.get() == "max_flow":

            algo = self.result["algorithm"]

            max_flow = self.result["max_flow"]

            tk.Label(
                card,
                text=f"{algo} Solution",
                font=("Segoe UI", 18, "bold"),
                bg="#e8f5e9",
                fg="#1b5e20"
            ).pack(pady=5)

            tk.Label(
                card,
                text=f"Maximum Flow = {max_flow}",
                font=("Segoe UI", 15, "bold"),
                bg="#e8f5e9",
                fg="#2e7d32"
            ).pack(pady=5)
        elif self.problem_type.get() == "mst":

            algo = self.result["algorithm"]
            total_weight = self.result["total_weight"]

            tk.Label(
                card,
                text=f"{algo} MST Solution",
                font=("Segoe UI", 18, "bold"),
                bg="#e8f5e9",
                fg="#1b5e20"
            ).pack(pady=5)

            tk.Label(
                card,
                text=f"Minimum Spanning Tree Weight = {total_weight}",
                font=("Segoe UI", 15, "bold"),
                bg="#e8f5e9",
                fg="#2e7d32"
            ).pack(pady=5)
        # ----------------------------------------------------
        # Detailed Results
        # ----------------------------------------------------

        detail_frame = tk.Frame(
            main,
            bg="white",
            bd=1,
            relief=tk.SOLID
        )

        detail_frame.pack(
            fill=tk.X,
            pady=10
        )

        # ----------------------------------------------------
        # Table Title
        # ----------------------------------------------------

        tk.Label(
            detail_frame,
            text="Detailed Results",
            font=("Segoe UI", 14, "bold"),
            bg="white",
            fg=self.HEADER_COLOR
        ).pack(anchor="w", padx=15, pady=10)

        # ----------------------------------------------------
        # Treeview
        # ----------------------------------------------------

        tree_frame = tk.Frame(
            detail_frame,
            bg="white"
        )

        tree_frame.pack(
            fill=tk.BOTH,
            expand=True,
            padx=15,
            pady=10
        )

        scrollbar = tk.Scrollbar(tree_frame)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        if self.problem_type.get() == "shortest_path":

            columns = ("Node", "Distance")

            tree = ttk.Treeview(
                tree_frame,
                columns=columns,
                show="headings",
                yscrollcommand=scrollbar.set,
                height=6
            )

            tree.heading("Node", text="Node")
            tree.heading("Distance", text="Distance From Source")

            tree.column("Node", width=180, anchor="center")
            tree.column("Distance", width=220, anchor="center")

            for node, dist in self.result["all_distances"].items():

                if dist == float("inf"):
                    dist = "∞"

                tree.insert(
                    "",
                    tk.END,
                    values=(node, dist)
                )

        elif self.problem_type.get() == "max_flow":

            columns = ("Path", "Flow")

            tree = ttk.Treeview(
                tree_frame,
                columns=columns,
                show="headings",
                yscrollcommand=scrollbar.set,
                height=10
            )

            tree.heading("Path", text="Augmenting Path")
            tree.heading("Flow", text="Flow Added")

            tree.column("Path", width=350, anchor="center")
            tree.column("Flow", width=150, anchor="center")

            for item in self.result["paths"]:

                path_text = " → ".join(item["path"])

                tree.insert(
                    "",
                    tk.END,
                    values=(path_text, item["flow"])
                )

        elif self.problem_type.get() == "mst":

            columns = ("Edge", "Weight")

            tree = ttk.Treeview(
                tree_frame,
                columns=columns,
                show="headings",
                yscrollcommand=scrollbar.set,
                height=10
            )

            tree.heading("Edge", text="Selected MST Edge")
            tree.heading("Weight", text="Weight")

            tree.column("Edge", width=350, anchor="center")
            tree.column("Weight", width=150, anchor="center")

            for u, v, w in self.result["mst_edges"]:

                tree.insert(
                    "",
                    tk.END,
                    values=(u + " - " + v, w)
                )

        tree.pack(fill=tk.BOTH, expand=True)

        scrollbar.config(command=tree.yview)

        # ----------------------------------------------------
        # Steps
        # ----------------------------------------------------

        steps_frame = tk.Frame(
            main,
            bg="white",
            bd=1,
            relief=tk.SOLID
        )

        steps_frame.pack(
            fill=tk.X,
            pady=10
        )

        tk.Label(
            steps_frame,
            text="Algorithm Steps",
            font=("Segoe UI", 14, "bold"),
            bg="white",
            fg=self.HEADER_COLOR
        ).pack(anchor="w", padx=15, pady=10)

        step_text = tk.Text(
            steps_frame,
            height=6,
            font=("Consolas", 10),
            bd=0
        )

        step_text.pack(
            fill=tk.BOTH,
            expand=True,
            padx=15,
            pady=10
        )

        if self.problem_type.get() == "shortest_path":

            for s in self.result["steps"]:
                step_text.insert(tk.END, s + "\n")

        elif self.problem_type.get() == "max_flow":

            for i, item in enumerate(self.result["paths"], start=1):

                txt = (
                    f"Iteration {i}: "
                    f"{' → '.join(item['path'])} "
                    f"| Flow = {item['flow']}"
                )

                step_text.insert(tk.END, txt + "\n")

        elif self.problem_type.get() == "mst":

            for s in self.result["steps"]:
                step_text.insert(tk.END, s + "\n")

        step_text.config(state=tk.DISABLED)

        # ----------------------------------------------------
        # Buttons
        # ----------------------------------------------------

        btn_frame = tk.Frame(
            main,
            bg=self.BG_COLOR
        )

        btn_frame.pack(pady=15)

        tk.Button(
            btn_frame,
            text="← Back",
            font=("Segoe UI", 11),
            bg="#757575",
            fg="white",
            padx=20,
            pady=8,
            command=lambda: self.notebook.select(self.page3)
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="Visualization →",
            font=("Segoe UI", 12, "bold"),
            bg=self.BUTTON_COLOR,
            fg="white",
            padx=25,
            pady=9,
            # command=self._create_page5
            command=self._go_to_visualization
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="Export TXT",
            font=("Segoe UI", 11, "bold"),
            bg="#455a64",
            fg="white",
            padx=18,
            pady=8,
            command=self._export_txt
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="Export Excel",
            font=("Segoe UI", 11, "bold"),
            bg="#2e7d32",
            fg="white",
            padx=18,
            pady=8,
            command=self._export_excel
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="Export PDF",
            font=("Segoe UI", 11, "bold"),
            bg="#c62828",
            fg="white",
            padx=18,
            pady=8,
            command=self._export_pdf
        ).pack(side=tk.LEFT, padx=10)


    def _go_to_visualization(self):

        self._create_page5()
        self.notebook.select(self.page5)

    # --------------------------------------------------------
    # Page 5 - Visualization
    # --------------------------------------------------------

    def _create_page5(self):

        for widget in self.page5.winfo_children():
            widget.destroy()

        main = tk.Frame(self.page5, bg=self.BG_COLOR)
        main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        tk.Label(
            main,
            text="Network Visualization",
            font=("Segoe UI", 22, "bold"),
            bg=self.BG_COLOR,
            fg=self.HEADER_COLOR
        ).pack(anchor="w", pady=(0, 15))

        info = tk.Frame(main, bg="#e3f2fd", bd=1, relief=tk.SOLID)
        info.pack(fill=tk.X, pady=10, ipady=8)

        if self.problem_type.get() == "shortest_path":
            info_text = "Highlighted path: " + " → ".join(self.result["path"])
        elif self.problem_type.get() == "max_flow":
            info_text = "Maximum Flow = " + str(self.result["max_flow"])
        elif self.problem_type.get() == "mst":
            info_text = "Minimum Spanning Tree Weight = " + str(self.result["total_weight"])
        else:
            info_text = "Visualization"

        tk.Label(
            info,
            text=info_text,
            font=("Segoe UI", 12, "bold"),
            bg="#e3f2fd",
            fg="#1565c0"
        ).pack()

        image_frame = tk.Frame(main, bg="white", bd=1, relief=tk.SOLID)
        image_frame.pack(fill=tk.BOTH, expand=True, pady=15)

        try:
            image_path = self._draw_graph_image()

            img = Image.open(image_path)
            img.thumbnail((1100, 520))

            self.graph_image = ImageTk.PhotoImage(img)

            image_label = tk.Label(
                image_frame,
                image=self.graph_image,
                bg="white"
            )
            image_label.pack(expand=True, pady=10)

        except Exception as e:
            tk.Label(
                image_frame,
                text="Could not draw graph:\n" + str(e),
                font=("Segoe UI", 12),
                bg="white",
                fg="#c62828"
            ).pack(expand=True)

        btn_frame = tk.Frame(main, bg=self.BG_COLOR)
        btn_frame.pack(pady=10)

        tk.Button(
            btn_frame,
            text="← Back to Results",
            font=("Segoe UI", 11),
            bg="#757575",
            fg="white",
            padx=20,
            pady=8,
            command=lambda: self.notebook.select(self.page4)
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="Save Graph PNG",
            font=("Segoe UI", 12, "bold"),
            bg=self.BUTTON_COLOR,
            fg="white",
            padx=25,
            pady=9,
            command=self._save_graph_png
        ).pack(side=tk.LEFT, padx=10)

    def _draw_graph_image(self):

        if self.problem_type.get() == "mst":
            G = nx.Graph()
        else:
            G = nx.DiGraph()

        for u, v, value in self.edges:
            G.add_edge(u, v, weight=value)

        pos = nx.spring_layout(G, seed=42)

        plt.figure(figsize=(11, 6))
        plt.axis("off")

        node_colors = []

        for node in G.nodes():
            if node == self.start_node:
                node_colors.append("#2e7d32")
            elif node == self.end_node:
                node_colors.append("#c62828")
            else:
                node_colors.append("#90caf9")

        nx.draw_networkx_nodes(
            G,
            pos,
            node_color=node_colors,
            node_size=1800,
            edgecolors="black",
            linewidths=1.2
        )

        nx.draw_networkx_labels(
            G,
            pos,
            font_size=11,
            font_weight="bold"
        )

        default_edges = list(G.edges())
        highlighted_edges = []

        if self.problem_type.get() == "shortest_path":
            path = self.result["path"]
            for i in range(len(path) - 1):
                highlighted_edges.append((path[i], path[i + 1]))

        elif self.problem_type.get() == "max_flow":
            for item in self.result["paths"]:
                path = item["path"]
                for i in range(len(path) - 1):
                    edge = (path[i], path[i + 1])
                    if edge not in highlighted_edges:
                        highlighted_edges.append(edge)
        elif self.problem_type.get() == "mst":
            for u, v, w in self.result["mst_edges"]:
                highlighted_edges.append((u, v))
                highlighted_edges.append((v, u))
        normal_edges = [
            e for e in default_edges if e not in highlighted_edges
        ]

        nx.draw_networkx_edges(
            G,
            pos,
            edgelist=normal_edges,
            edge_color="#9e9e9e",
            arrows=(self.problem_type.get() != "mst"),
            arrowsize=20,
            width=1.8
        )

        nx.draw_networkx_edges(
            G,
            pos,
            edgelist=highlighted_edges,
            edge_color="#f57c00",
            arrows=(self.problem_type.get() != "mst"),
            arrowsize=24,
            width=3.2
        )

        edge_labels = {}

        for u, v, value in self.edges:
            if self.problem_type.get() == "shortest_path":
                edge_labels[(u, v)] = str(value)
            elif self.problem_type.get() == "max_flow":
                edge_labels[(u, v)] = "cap=" + str(value)
            elif self.problem_type.get() == "mst":
                edge_labels[(u, v)] = str(value)
        nx.draw_networkx_edge_labels(
            G,
            pos,
            edge_labels=edge_labels,
            font_size=9,
            label_pos=0.5,
            bbox=dict(
                boxstyle="round,pad=0.2",
                fc="white",
                ec="none"
            )
        )

        title = ""

        if self.problem_type.get() == "shortest_path":
            title = (
                "Shortest Path: "
                + " → ".join(self.result["path"])
                + " | Distance = "
                + str(self.result["distance"])
            )

        elif self.problem_type.get() == "max_flow":
            title = "Maximum Flow = " + str(self.result["max_flow"])
        elif self.problem_type.get() == "mst":
            title = "MST Total Weight = " + str(self.result["total_weight"])
        plt.title(title, fontsize=14, fontweight="bold")

        output_path = os.path.join(
            os.getcwd(),
            "network_graph_preview.png"
        )

        plt.tight_layout()
        plt.savefig(output_path, dpi=160)
        plt.close()

        self.current_graph_path = output_path

        return output_path

    def _save_graph_png(self):

        try:
            if not hasattr(self, "current_graph_path"):
                self._draw_graph_image()

            filename = filedialog.asksaveasfilename(
                initialdir=os.getcwd(),
                initialfile="network_graph.png",
                defaultextension=".png",
                filetypes=[
                    ("PNG Image", "*.png"),
                    ("All files", "*.*")
                ],
                title="Save Network Graph"
            )

            if not filename:
                return

            import shutil
            shutil.copyfile(self.current_graph_path, filename)

            messagebox.showinfo(
                "Saved",
                "Graph image saved to:\n" + filename
            )

        except Exception as e:
            messagebox.showerror(
                "Save Error",
                str(e)
            )


    # --------------------------------------------------------
    # Export TXT
    # --------------------------------------------------------

    def _export_txt(self):

        try:

            filename = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text File", "*.txt")]
            )

            if not filename:
                return

            with open(filename, "w", encoding="utf-8") as f:

                f.write("NETWORK OPTIMIZATION REPORT\n")
                f.write("=" * 50 + "\n\n")

                f.write(
                    "Problem Type: "
                    + self.problem_type.get()
                    + "\n"
                )

                f.write(
                    "Algorithm: "
                    + self.result["algorithm"]
                    + "\n\n"
                )

                if self.problem_type.get() == "shortest_path":

                    f.write(
                        "Shortest Distance = "
                        + str(self.result["distance"])
                        + "\n"
                    )

                    f.write(
                        "Path = "
                        + " -> ".join(self.result["path"])
                        + "\n\n"
                    )

                    f.write("Distances:\n")

                    for node, dist in self.result["all_distances"].items():

                        f.write(f"{node}: {dist}\n")

                elif self.problem_type.get() == "max_flow":

                    f.write(
                        "Maximum Flow = "
                        + str(self.result["max_flow"])
                        + "\n\n"
                    )

                    f.write("Augmenting Paths:\n")

                    for item in self.result["paths"]:

                        f.write(
                            " -> ".join(item["path"])
                            + f" | Flow = {item['flow']}\n"
                        )
                elif self.problem_type.get() == "mst":

                    f.write(
                        "Minimum Spanning Tree Total Weight = "
                        + str(self.result["total_weight"])
                        + "\n\n"
                    )

                    f.write("Selected MST Edges:\n")

                    for u, v, w in self.result["mst_edges"]:
                        f.write(f"{u} - {v} | Weight = {w}\n")
            messagebox.showinfo(
                "Export Complete",
                "TXT report saved successfully."
            )

        except Exception as e:

            messagebox.showerror(
                "Export Error",
                str(e)
            )

    # --------------------------------------------------------
    # Export Excel
    # --------------------------------------------------------

    def _export_excel(self):

        try:

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel File", "*.xlsx")]
            )

            if not filename:
                return

            wb = Workbook()

            ws = wb.active

            ws.title = "Network Results"

            title_font = Font(
                bold=True,
                size=14,
                color="FFFFFF"
            )

            header_fill = PatternFill(
                start_color="1565C0",
                end_color="1565C0",
                fill_type="solid"
            )

            center = Alignment(
                horizontal="center"
            )

            ws["A1"] = "Network Optimization Report"

            ws["A1"].font = title_font
            ws["A1"].fill = header_fill

            row = 3

            ws[f"A{row}"] = "Problem Type"
            ws[f"B{row}"] = self.problem_type.get()

            row += 1

            ws[f"A{row}"] = "Algorithm"
            ws[f"B{row}"] = self.result["algorithm"]

            row += 2

            if self.problem_type.get() == "shortest_path":

                ws[f"A{row}"] = "Shortest Distance"
                ws[f"B{row}"] = self.result["distance"]

                row += 1

                ws[f"A{row}"] = "Path"
                ws[f"B{row}"] = " -> ".join(
                    self.result["path"]
                )

                row += 2

                ws[f"A{row}"] = "Node"
                ws[f"B{row}"] = "Distance"

                ws[f"A{row}"].fill = header_fill
                ws[f"B{row}"].fill = header_fill

                row += 1

                for node, dist in self.result["all_distances"].items():

                    ws[f"A{row}"] = node
                    ws[f"B{row}"] = dist

                    row += 1

            elif self.problem_type.get() == "max_flow":

                ws[f"A{row}"] = "Maximum Flow"
                ws[f"B{row}"] = self.result["max_flow"]

                row += 2

                ws[f"A{row}"] = "Path"
                ws[f"B{row}"] = "Flow"

                ws[f"A{row}"].fill = header_fill
                ws[f"B{row}"].fill = header_fill

                row += 1

                for item in self.result["paths"]:

                    ws[f"A{row}"] = " -> ".join(item["path"])
                    ws[f"B{row}"] = item["flow"]

                    row += 1
            elif self.problem_type.get() == "mst":

                ws[f"A{row}"] = "MST Total Weight"
                ws[f"B{row}"] = self.result["total_weight"]

                row += 2

                ws[f"A{row}"] = "Edge"
                ws[f"B{row}"] = "Weight"

                ws[f"A{row}"].fill = header_fill
                ws[f"B{row}"].fill = header_fill

                row += 1

                for u, v, w in self.result["mst_edges"]:
                    ws[f"A{row}"] = u + " - " + v
                    ws[f"B{row}"] = w
                    row += 1
            for col in ["A", "B", "C", "D"]:

                ws.column_dimensions[col].width = 28

            try:
                graph_path = self._draw_graph_image()

                img = XLImage(graph_path)

                img.width = 700
                img.height = 400

                ws.add_image(img, "D3")

            except Exception as e:
                print("Could not add graph to Excel:", e)

            wb.save(filename)

            messagebox.showinfo(
                "Export Complete",
                "Excel report saved successfully."
            )

        except Exception as e:

            messagebox.showerror(
                "Export Error",
                str(e)
            )

    # --------------------------------------------------------
    # Export PDF
    # --------------------------------------------------------

    def _export_pdf(self):

        try:

            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF File", "*.pdf")]
            )

            if not filename:
                return

            doc = SimpleDocTemplate(
                filename,
                pagesize=A4
            )

            styles = getSampleStyleSheet()

            elements = []

            title = Paragraph(
                "<b>Network Optimization Report</b>",
                styles["Title"]
            )

            elements.append(title)
            elements.append(Spacer(1, 20))

            info = f"""
            <b>Problem Type:</b> {self.problem_type.get()}<br/>
            <b>Algorithm:</b> {self.result["algorithm"]}<br/>
            """

            elements.append(
                Paragraph(info, styles["BodyText"])
            )
            if self.problem_type.get() == "shortest_path":
                explanation = (
                    "This report analyzes a shortest path network problem. "
                    "The goal is to find the minimum total cost or distance from the selected "
                    "start node to the selected end node."
                )
            elif self.problem_type.get() == "max_flow":
                explanation = (
                    "This report analyzes a maximum flow network problem. "
                    "The goal is to send the largest possible amount of flow from the source "
                    "node to the sink node without violating arc capacities."
                )
            elif self.problem_type.get() == "mst":
                explanation = (
                    "This report analyzes a Minimum Spanning Tree problem. "
                    "The goal is to connect all nodes with minimum total edge weight without creating cycles."
                )
            else:
                explanation = "Network optimization report."

            elements.append(Spacer(1, 10))
            elements.append(Paragraph(explanation, styles["BodyText"]))

            elements.append(Spacer(1, 20))

            data = []

            if self.problem_type.get() == "shortest_path":

                data.append(["Metric", "Value"])

                data.append([
                    "Shortest Distance",
                    str(self.result["distance"])
                ])

                data.append([
                    "Path",
                    " -> ".join(self.result["path"])
                ])

            elif self.problem_type.get() == "max_flow":

                data.append(["Metric", "Value"])

                data.append([
                    "Maximum Flow",
                    str(self.result["max_flow"])
                ])
            elif self.problem_type.get() == "mst":

                data.append(["Metric", "Value"])

                data.append([
                    "MST Total Weight",
                    str(self.result["total_weight"])
                ])

                for u, v, w in self.result["mst_edges"]:
                    data.append([
                        u + " - " + v,
                        str(w)
                    ])
            table = Table(data, colWidths=[200, 280])

            table.setStyle(TableStyle([

                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),

                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

                ("GRID", (0, 0), (-1, -1), 1, colors.black),

                ("ALIGN", (0, 0), (-1, -1), "CENTER"),

                ("BOTTOMPADDING", (0, 0), (-1, 0), 10)

            ]))

            elements.append(table)
            elements.append(Spacer(1, 20))

            try:
                graph_path = self._draw_graph_image()
                elements.append(Paragraph("<b>Network Graph</b>", styles["Heading2"]))
                elements.append(Spacer(1, 10))
                elements.append(RLImage(graph_path, width=480, height=260))
            except Exception as graph_error:
                elements.append(Paragraph("Graph could not be added: " + str(graph_error), styles["BodyText"]))

            doc.build(elements)

            messagebox.showinfo(
                "Export Complete",
                "PDF report saved successfully."
            )

        except Exception as e:

            messagebox.showerror(
                "Export Error",
                str(e)
            )

# ============================================================
# Main
# ============================================================

if __name__ == "__main__":

    root = tk.Tk()

    app = NetworkGUI(root)

    root.mainloop()