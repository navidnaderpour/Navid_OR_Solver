import tkinter as tk
from tkinter import ttk, messagebox


# ============================================================
# Transport Problem Solver
# ============================================================

class TransportSolver:
    """
    Solves the Transportation Problem using NWC, Least Cost, VAM, and MODI methods.
    
    Input:
        supply:  list of supply amounts at origins       e.g. [10, 15, 20]
        demand:  list of demand amounts at destinations   e.g. [8, 12, 10, 15]
        costs:   cost matrix                              e.g. [[3,7,2,5],[4,1,8,6],[9,3,4,2]]
    """
    
    def __init__(self, supply, demand, costs):
        self.original_supply = supply[:]
        self.original_demand = demand[:]
        self.original_costs = [row[:] for row in costs]
        
        self.supply = supply[:]
        self.demand = demand[:]
        self.costs = [row[:] for row in costs]
        
        self.num_supply = len(supply)
        self.num_demand = len(demand)
        self.is_balanced_flag = False
        self.solution = None
        self.total_cost = None
        self.method_used = ""
        self.iterations = []
        
        self._check_balance()
    
    def _check_balance(self):
        """Check if total supply equals total demand"""
        total_supply = sum(self.supply)
        total_demand = sum(self.demand)
        
        if total_supply == total_demand:
            self.is_balanced_flag = True
            return True
        else:
            self.is_balanced_flag = False
            return False
    
    def _balance_problem(self):
        """If unbalanced, add a dummy origin or destination with zero cost"""
        total_supply = sum(self.supply)
        total_demand = sum(self.demand)
        
        if total_supply == total_demand:
            return
        
        if total_supply < total_demand:
            diff = total_demand - total_supply
            self.supply.append(diff)
            self.costs.append([0] * self.num_demand)
            self.num_supply += 1
            print(f"NOTE: Unbalanced problem - supply shortage of {diff}. Added dummy origin S{self.num_supply} with zero cost.")
        else:
            diff = total_supply - total_demand
            self.demand.append(diff)
            for row in self.costs:
                row.append(0)
            self.num_demand += 1
            print(f"NOTE: Unbalanced problem - demand shortage of {diff}. Added dummy destination D{self.num_demand} with zero cost.")
    
    # ============================================================
    # Method 1: North-West Corner Method (NWC)
    # ============================================================
    
    def nwc(self):
        """
        North-West Corner Method (NWC)
        
        Simple: start at cell (0,0), allocate as much as possible,
        then move right or down and repeat.
        """
        if not self._check_balance():
            self._balance_problem()
        
        supply = self.supply[:]
        demand = self.demand[:]
        costs = [row[:] for row in self.costs]
        
        solution_matrix = [[0] * self.num_demand for _ in range(self.num_supply)]
        
        i = 0
        j = 0
        
        total_cost = 0
        steps = []
        step_num = 1
        
        while i < self.num_supply and j < self.num_demand:
            amount = min(supply[i], demand[j])
            
            solution_matrix[i][j] = amount
            cost_value = costs[i][j] * amount
            total_cost += cost_value
            
            steps.append({
                'step': step_num,
                'position': f"S{i+1} -> D{j+1}",
                'amount': amount,
                'cost': costs[i][j],
                'total_step_cost': cost_value,
            })
            
            supply[i] -= amount
            demand[j] -= amount
            
            step_num += 1
            
            if supply[i] == 0 and demand[j] == 0:
                i += 1
                j += 1
            elif supply[i] == 0:
                i += 1
            else:
                j += 1
        
        self.solution = solution_matrix
        self.total_cost = total_cost
        self.method_used = "NWC (North-West Corner)"
        self.iterations = steps
        
        return solution_matrix, total_cost, steps
    
    # ============================================================
    # Method 2: Least Cost Method (LCM)
    # ============================================================
    
    def least_cost(self):
        """
        Least Cost Method (LCM)
        
        Better than NWC: always pick the cheapest available cell.
        """
        if not self._check_balance():
            self._balance_problem()
        
        supply = self.supply[:]
        demand = self.demand[:]
        costs = [row[:] for row in self.costs]
        
        available_rows = [True] * self.num_supply
        available_cols = [True] * self.num_demand
        
        solution_matrix = [[0] * self.num_demand for _ in range(self.num_supply)]
        
        total_cost = 0
        steps = []
        step_num = 1
        
        while any(available_rows) and any(available_cols):
            min_cost = float('inf')
            min_i = -1
            min_j = -1
            
            for i in range(self.num_supply):
                if not available_rows[i]:
                    continue
                for j in range(self.num_demand):
                    if not available_cols[j]:
                        continue
                    if costs[i][j] < min_cost:
                        min_cost = costs[i][j]
                        min_i = i
                        min_j = j
            
            if min_i == -1 or min_j == -1:
                break
            
            i, j = min_i, min_j
            amount = min(supply[i], demand[j])
            
            solution_matrix[i][j] = amount
            cost_value = costs[i][j] * amount
            total_cost += cost_value
            
            steps.append({
                'step': step_num,
                'position': f"S{i+1} -> D{j+1}",
                'amount': amount,
                'cost': costs[i][j],
                'total_step_cost': cost_value,
            })
            
            supply[i] -= amount
            demand[j] -= amount
            step_num += 1
            
            if supply[i] == 0:
                available_rows[i] = False
            if demand[j] == 0:
                available_cols[j] = False
        
        self.solution = solution_matrix
        self.total_cost = total_cost
        self.method_used = "Least Cost (LCM)"
        self.iterations = steps
        
        return solution_matrix, total_cost, steps
    
    # ============================================================
    # Method 3: Vogel's Approximation Method (VAM)
    # ============================================================
    
    def vogel(self):
        """
        Vogel's Approximation Method (VAM)
        
        Best initial solution. Uses row/column penalties.
        """
        if not self._check_balance():
            self._balance_problem()
        
        supply = self.supply[:]
        demand = self.demand[:]
        costs = [row[:] for row in self.costs]
        
        available_rows = [True] * self.num_supply
        available_cols = [True] * self.num_demand
        
        solution_matrix = [[0] * self.num_demand for _ in range(self.num_supply)]
        
        total_cost = 0
        steps = []
        step_num = 1
        
        while any(available_rows) and any(available_cols):
            row_penalties = []
            for i in range(self.num_supply):
                if not available_rows[i]:
                    row_penalties.append(-1)
                    continue
                avail_costs = [costs[i][j] for j in range(self.num_demand) if available_cols[j]]
                if len(avail_costs) >= 2:
                    sorted_costs = sorted(avail_costs)
                    penalty = sorted_costs[1] - sorted_costs[0]
                elif len(avail_costs) == 1:
                    penalty = avail_costs[0]
                else:
                    penalty = -1
                row_penalties.append(penalty)
            
            col_penalties = []
            for j in range(self.num_demand):
                if not available_cols[j]:
                    col_penalties.append(-1)
                    continue
                avail_costs = [costs[i][j] for i in range(self.num_supply) if available_rows[i]]
                if len(avail_costs) >= 2:
                    sorted_costs = sorted(avail_costs)
                    penalty = sorted_costs[1] - sorted_costs[0]
                elif len(avail_costs) == 1:
                    penalty = avail_costs[0]
                else:
                    penalty = -1
                col_penalties.append(penalty)
            
            max_row_penalty = max(row_penalties)
            max_col_penalty = max(col_penalties)
            
            if max_row_penalty >= max_col_penalty:
                i = row_penalties.index(max_row_penalty)
                min_cost = float('inf')
                j = -1
                for col in range(self.num_demand):
                    if available_cols[col] and costs[i][col] < min_cost:
                        min_cost = costs[i][col]
                        j = col
            else:
                j = col_penalties.index(max_col_penalty)
                min_cost = float('inf')
                i = -1
                for row in range(self.num_supply):
                    if available_rows[row] and costs[row][j] < min_cost:
                        min_cost = costs[row][j]
                        i = row
            
            if i == -1 or j == -1:
                break
            
            amount = min(supply[i], demand[j])
            
            solution_matrix[i][j] = amount
            cost_value = costs[i][j] * amount
            total_cost += cost_value
            
            steps.append({
                'step': step_num,
                'position': f"S{i+1} -> D{j+1}",
                'amount': amount,
                'cost': costs[i][j],
                'total_step_cost': cost_value,
            })
            
            supply[i] -= amount
            demand[j] -= amount
            step_num += 1
            
            if supply[i] == 0:
                available_rows[i] = False
            if demand[j] == 0:
                available_cols[j] = False
        
        self.solution = solution_matrix
        self.total_cost = total_cost
        self.method_used = "VAM (Vogel's Approximation)"
        self.iterations = steps
        
        return solution_matrix, total_cost, steps
    
    # ============================================================
    # Method 4: MODI (Modified Distribution) - Optimization
    # ============================================================
    
    def _get_loop(self, solution_matrix, start_i, start_j, basic_cells=None):
        """
        Find a closed MODI loop for the entering cell (start_i, start_j).

        The loop must:
        - start from the entering non-basic cell
        - move only horizontally or vertically
        - turn only at basic cells
        - return to the starting cell
        """

        rows = self.num_supply
        cols = self.num_demand

        if basic_cells is None:
            basic_cells = []
            for i in range(rows):
                for j in range(cols):
                    if solution_matrix[i][j] > 0:
                        basic_cells.append((i, j))

        start = (start_i, start_j)
        allowed_cells = set(basic_cells)
        allowed_cells.add(start)

        def dfs(current, path, move_direction):
            i, j = current

            # If we returned to start and have a valid loop
            if current == start and len(path) >= 4:
                return path

            # Horizontal move: same row, different column
            if move_direction != "horizontal":
                for next_j in range(cols):
                    nxt = (i, next_j)
                    if nxt == current:
                        continue
                    if nxt not in allowed_cells:
                        continue
                    if nxt in path and nxt != start:
                        continue
                    if nxt == start and len(path) < 4:
                        continue

                    result = dfs(nxt, path + [nxt], "horizontal")
                    if result:
                        return result

            # Vertical move: same column, different row
            if move_direction != "vertical":
                for next_i in range(rows):
                    nxt = (next_i, j)
                    if nxt == current:
                        continue
                    if nxt not in allowed_cells:
                        continue
                    if nxt in path and nxt != start:
                        continue
                    if nxt == start and len(path) < 4:
                        continue

                    result = dfs(nxt, path + [nxt], "vertical")
                    if result:
                        return result

            return None

        loop = dfs(start, [start], None)

        if loop and loop[-1] == start:
            return loop[:-1]

        return None
    
    def modi(self, initial_method='vogel'):
        """
        MODI (Modified Distribution) Method
        
        Takes an initial solution and optimizes it.
        
        Parameters:
            initial_method: 'nwc', 'least_cost', or 'vogel' (default: 'vogel')
        """
        # Step 0: Get initial solution
        if initial_method == 'nwc':
            self.nwc()
        elif initial_method == 'least_cost':
            self.least_cost()
        else:
            self.vogel()
        
        solution_matrix = [row[:] for row in self.solution]
        supply = self.supply[:]
        demand = self.demand[:]
        costs = [row[:] for row in self.costs]
        
        rows = self.num_supply
        cols = self.num_demand
        
        total_cost = sum(solution_matrix[i][j] * costs[i][j] 
                         for i in range(rows) for j in range(cols))
        
        optimization_steps = []
        iteration = 1
        
        while True:
            u = [None] * rows
            v = [None] * cols
            u[0] = 0
            
            basic_cells = [(i, j) for i in range(rows) for j in range(cols)
                           if solution_matrix[i][j] > 0]

            # Handle degeneracy without changing real allocation values
            if len(basic_cells) < rows + cols - 1:
                print("  NOTE: Degenerate solution. Adding artificial basic cells.")

                for i in range(rows):
                    for j in range(cols):
                        if (i, j) not in basic_cells:
                          
                            basic_cells.append((i, j))

                            if len(basic_cells) >= rows + cols - 1:
                                break
                    if len(basic_cells) >= rows + cols - 1:
                        break
            
            # Solve u_i + v_j = c_ij for basic cells
            changed = True
            while changed:
                changed = False
                for i, j in basic_cells:
                    if u[i] is not None and v[j] is None:
                        v[j] = costs[i][j] - u[i]
                        changed = True
                    elif v[j] is not None and u[i] is None:
                        u[i] = costs[i][j] - v[j]
                        changed = True
            
            for i in range(rows):
                if u[i] is None:
                    u[i] = 0
            for j in range(cols):
                if v[j] is None:
                    v[j] = 0
            
            # Calculate reduced costs
            reduced_costs = {}
            for i in range(rows):
                for j in range(cols):
                    if (i, j) not in basic_cells:
                        d_ij = costs[i][j] - u[i] - v[j]
                        reduced_costs[(i, j)] = d_ij
            
            negative_cells = [(i, j) for (i, j), d in reduced_costs.items() if d < -0.0001]
            
            iter_info = {
                'iteration': iteration,
                'u_values': u[:],
                'v_values': v[:],
                'reduced_costs': dict(reduced_costs),
                'solution': [row[:] for row in solution_matrix],
                'total_cost': total_cost,
                'negative_cells': negative_cells[:]
            }
            optimization_steps.append(iter_info)
            
            if not negative_cells:
                print(f"  MODI converged after {iteration} iteration(s).")
                break
            
            pivot_cell = min(negative_cells, key=lambda x: reduced_costs[x])
            pivot_i, pivot_j = pivot_cell
            print(f"  Iteration {iteration}: Pivot on S{pivot_i+1} -> D{pivot_j+1} (d = {reduced_costs[pivot_cell]:.2f})")
            
            loop = self._get_loop(solution_matrix, pivot_i, pivot_j, basic_cells)
            
            if loop is None or len(loop) < 4:
                print(f"  WARNING: Could not find loop. Stopping MODI.")
                break
            
            # Find theta
            theta = float('inf')
            for k in range(1, len(loop), 2):
                i, j = loop[k]
                if solution_matrix[i][j] > 0:
                    theta = min(theta, solution_matrix[i][j])
            if theta == float('inf'):
                print("  WARNING: Could not find theta. Stopping MODI.")
                break
           
            
            # Apply theta
            for k, (i, j) in enumerate(loop):
                if k % 2 == 0:
                    solution_matrix[i][j] += theta
                else:
                    solution_matrix[i][j] -= theta
                    if abs(solution_matrix[i][j]) < 1e-9:
                        solution_matrix[i][j] = 0
            
            total_cost = sum(solution_matrix[i][j] * costs[i][j] 
                             for i in range(rows) for j in range(cols))
            
            iteration += 1
            
            if iteration > 20:
                print("  WARNING: Max iterations reached.")
                break
        
        self.solution = solution_matrix
        self.total_cost = total_cost
        self.method_used = f"MODI (from {initial_method.title()})"
        self.iterations = optimization_steps
        
        return solution_matrix, total_cost, optimization_steps
    
    # ============================================================
    # Display Solution
    # ============================================================
    
    def print_solution(self):
        """Pretty-print the solution to console"""
        if self.solution is None:
            print("ERROR: Problem not solved yet!")
            return
        
        print(f"\n{'='*60}")
        print(f"TRANSPORTATION PROBLEM - Method: {self.method_used}")
        print(f"{'='*60}")
        
        print(f"\n{'':>10}", end="")
        for j in range(self.num_demand):
            print(f"{f'D{j+1}':>8}", end="")
        print(f"{'Supply':>8}")
        print("-" * (10 + 8 * self.num_demand + 8))
        
        for i in range(self.num_supply):
            print(f"{f'S{i+1}':>10}", end="")
            for j in range(self.num_demand):
                if self.solution[i][j] > 0:
                    if self.solution[i][j] == int(self.solution[i][j]):
                        print(f"{int(self.solution[i][j]):>8}", end="")
                    else:
                        print(f"{self.solution[i][j]:>8.1f}", end="")
                else:
                    print(f"{'·':>8}", end="")
            print(f"{self.supply[i]:>8}")
        
        print(f"{'Demand':>10}", end="")
        for j in range(self.num_demand):
            print(f"{self.demand[j]:>8}", end="")
        print()
        
        print(f"\n{'─'*60}")
        print(f"TOTAL COST: {self.total_cost}")
        print(f"{'─'*60}")
        
        # Check iterations format
        if self.iterations and isinstance(self.iterations[0], dict):
            if 'step' in self.iterations[0]:
                print(f"\nSolution Steps:")
                print(f"{'Step':<8} {'Route':<12} {'Qty':<8} {'Unit Cost':<12} {'Step Cost':<12}")
                print("-" * 52)
                for step in self.iterations:
                    print(f"{step['step']:<8} {step['position']:<12} {step['amount']:<8} {step['cost']:<12} {step['total_step_cost']:<12}")
            elif 'iteration' in self.iterations[0]:
                print(f"\nMODI Optimization Steps:")
                print("-" * 60)
                for it in self.iterations:
                    neg = it['negative_cells']
                    print(f"  Iteration {it['iteration']}: Total Cost = {it['total_cost']}, Negative cells: {len(neg)}")
                    if neg:
                        for ni, nj in neg[:3]:
                            print(f"    S{ni+1}->D{nj+1}: d = {it['reduced_costs'].get((ni,nj), 0):.2f}")
        else:
            print(f"\nTotal allocations: {sum(1 for row in self.solution for v in row if v > 0)}")
    
    def get_solution_summary(self):
        """Return summary dict for GUI display"""
        if self.solution is None:
            return None
        
        summary = []
        for i in range(self.num_supply):
            for j in range(self.num_demand):
                if self.solution[i][j] > 0:
                    summary.append({
                        'from': f"S{i+1}",
                        'to': f"D{j+1}",
                        'amount': self.solution[i][j],
                        'unit_cost': self.costs[i][j],
                        'total_cost': self.costs[i][j] * self.solution[i][j]
                    })
        
        return {
            'method': self.method_used,
            'total_cost': self.total_cost,
            'assignments': summary,
            'steps': self.iterations,
            'is_balanced': self.is_balanced_flag,
            'num_supply': self.num_supply,
            'num_demand': self.num_demand
        }





# ============================================================
# GUI - Transportation Problem Solver
# ============================================================

class TransportGUI:
    """GUI for the Transportation Problem Solver"""
    
    BG_COLOR = "#f0f4f8"
    HEADER_COLOR = "#1a237e"
    BUTTON_COLOR = "#283593"
    SUCCESS_COLOR = "#2e7d32"
    TABLE_HEADER_COLOR = "#e8eaf6"
    ROW_ALT_COLOR = "#f5f5f5"
    
    def __init__(self, root):
        self.root = root
        self.root.title("Transportation Problem Solver")
        self.root.geometry("1024x700")
        self.root.configure(bg=self.BG_COLOR)
        
        self.num_supply = tk.IntVar(value=3)
        self.num_demand = tk.IntVar(value=4)
        self.supply_entries = []
        self.demand_entries = []
        self.cost_entries = []
        self.solver = None
        self.results = {}
        
        self._setup_styles()
        self._create_header()
        self._create_notebook()
        self._create_page1()
    
    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TNotebook', background=self.BG_COLOR, borderwidth=0)
        style.configure('TNotebook.Tab', padding=[12, 4], font=('Segoe UI', 10))
        style.map('TNotebook.Tab', background=[('selected', '#283593')], foreground=[('selected', 'white')])
        style.configure('Header.TLabel', font=('Segoe UI', 14, 'bold'), foreground=self.HEADER_COLOR)
        style.configure('SubHeader.TLabel', font=('Segoe UI', 11), foreground='#424242')
        style.configure('Result.TLabel', font=('Segoe UI', 12, 'bold'), foreground=self.SUCCESS_COLOR)
        style.configure('TButton', font=('Segoe UI', 10, 'bold'), padding=8)
    
    def _create_header(self):
        header = tk.Frame(self.root, bg=self.HEADER_COLOR, height=56)
        header.pack(fill=tk.X, side=tk.TOP)
        header.pack_propagate(False)
        tk.Label(header, text="Transportation Problem Solver", 
                font=('Segoe UI', 16, 'bold'), fg='white', bg=self.HEADER_COLOR).pack(side=tk.LEFT, padx=20, pady=12)
        tk.Label(header, text="Prepared by Navid Naderpour, navid.naderpour@gmail.com",
                font=('Segoe UI', 9), fg='#bbdefb', bg=self.HEADER_COLOR).pack(side=tk.RIGHT, padx=20, pady=16)
    
    def _create_notebook(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.page1 = tk.Frame(self.notebook, bg=self.BG_COLOR)
        self.page2 = tk.Frame(self.notebook, bg=self.BG_COLOR)
        self.page3 = tk.Frame(self.notebook, bg=self.BG_COLOR)
        self.page4 = tk.Frame(self.notebook, bg=self.BG_COLOR)
        self.page5 = tk.Frame(self.notebook, bg=self.BG_COLOR)
        self.notebook.add(self.page1, text=" Page 1: Setup  ")
        self.notebook.add(self.page2, text=" Page 2: Data   ")
        self.notebook.add(self.page3, text=" Page 3: Solve  ")
        self.notebook.add(self.page4, text=" Page 4: Result ")
        self.notebook.add(self.page5, text=" Page 5: Compare")
    
    # ============================
    # PAGE 1
    # ============================
    
    def _create_page1(self):
        frame = tk.Frame(self.page1, bg=self.BG_COLOR)
        frame.pack(expand=True, pady=60)
        tk.Label(frame, text="Problem Setup", font=('Segoe UI', 18, 'bold'),
                bg=self.BG_COLOR, fg=self.HEADER_COLOR).pack(pady=(0, 30))
        
        row1 = tk.Frame(frame, bg=self.BG_COLOR)
        row1.pack(pady=10)
        tk.Label(row1, text="Number of Origins (Factories / Warehouses):", 
                font=('Segoe UI', 12), bg=self.BG_COLOR).pack(side=tk.LEFT, padx=10)
        spin_s = tk.Spinbox(row1, from_=2, to=20, textvariable=self.num_supply, 
                          width=5, font=('Segoe UI', 12), justify=tk.CENTER)
        spin_s.pack(side=tk.LEFT, padx=10)
        
        row2 = tk.Frame(frame, bg=self.BG_COLOR)
        row2.pack(pady=10)
        tk.Label(row2, text="Number of Destinations (Customers / Stores):", 
                font=('Segoe UI', 12), bg=self.BG_COLOR).pack(side=tk.LEFT, padx=10)
        spin_d = tk.Spinbox(row2, from_=2, to=20, textvariable=self.num_demand, 
                          width=5, font=('Segoe UI', 12), justify=tk.CENTER)
        spin_d.pack(side=tk.LEFT, padx=10)
        
        info_frame = tk.Frame(frame, bg='#e3f2fd', bd=1, relief=tk.SOLID)
        info_frame.pack(pady=30, ipadx=20, ipady=10)
        tk.Label(info_frame, text="Example: 3 factories shipping to 4 warehouses",
                font=('Segoe UI', 10), bg='#e3f2fd', fg='#1565c0').pack()
        
        tk.Button(frame, text="NEXT  ->  Enter Data", font=('Segoe UI', 12, 'bold'),
                 bg=self.BUTTON_COLOR, fg='white', padx=30, pady=8,
                 command=self._go_to_page2).pack(pady=20)
    
    def _go_to_page2(self):
        self._create_page2_data()
        self.notebook.select(self.page2)
    
    # ============================
    # PAGE 2
    # ============================
    
    def _create_page2_data(self):
        for widget in self.page2.winfo_children():
            widget.destroy()
        
        main_frame = tk.Frame(self.page2, bg=self.BG_COLOR)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tk.Label(main_frame, text="Data Entry", font=('Segoe UI', 16, 'bold'),
                bg=self.BG_COLOR, fg=self.HEADER_COLOR).pack(anchor=tk.W, pady=(0, 10))
        
        canvas = tk.Canvas(main_frame, bg=self.BG_COLOR, highlightthickness=0)
        scrollbar = tk.Scrollbar(main_frame, orient=tk.VERTICAL, command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=self.BG_COLOR)
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor=tk.NW)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        ns = self.num_supply.get()
        nd = self.num_demand.get()
        self.supply_entries = []
        self.demand_entries = []
        self.cost_entries = [[None] * nd for _ in range(ns)]
        
        # Supply
        tk.Label(scroll_frame, text="Supply (Origins)", font=('Segoe UI', 12, 'bold'),
                bg=self.BG_COLOR, fg=self.HEADER_COLOR).pack(anchor=tk.W, pady=(10, 5))
        sf = tk.Frame(scroll_frame, bg=self.BG_COLOR)
        sf.pack(fill=tk.X, pady=5)
        for i in range(ns):
            tk.Label(sf, text="  S" + str(i+1) + ":", font=('Segoe UI', 11),
                    bg=self.BG_COLOR, width=6, anchor=tk.W).grid(row=i, column=0, padx=5, pady=3)
            e = tk.Entry(sf, width=12, font=('Segoe UI', 11), justify=tk.CENTER, relief=tk.SOLID, bd=1)
            e.grid(row=i, column=1, padx=5, pady=3)
            e.insert(0, "10")
            self.supply_entries.append(e)
        
        # Demand
        tk.Label(scroll_frame, text="Demand (Destinations)", font=('Segoe UI', 12, 'bold'),
                bg=self.BG_COLOR, fg=self.HEADER_COLOR).pack(anchor=tk.W, pady=(15, 5))
        df = tk.Frame(scroll_frame, bg=self.BG_COLOR)
        df.pack(fill=tk.X, pady=5)
        for j in range(nd):
            tk.Label(df, text="  D" + str(j+1) + ":", font=('Segoe UI', 11),
                    bg=self.BG_COLOR, width=6, anchor=tk.W).grid(row=0, column=j*2, padx=5, pady=3)
            e = tk.Entry(df, width=10, font=('Segoe UI', 11), justify=tk.CENTER, relief=tk.SOLID, bd=1)
            e.grid(row=1, column=j*2, padx=5, pady=3)
            e.insert(0, "8")
            self.demand_entries.append(e)
        
        # Cost matrix
        tk.Label(scroll_frame, text="Cost Matrix (Row = Origin, Col = Destination)", 
                font=('Segoe UI', 12, 'bold'), bg=self.BG_COLOR, fg=self.HEADER_COLOR).pack(anchor=tk.W, pady=(15, 5))
        cf = tk.Frame(scroll_frame, bg=self.BG_COLOR)
        cf.pack(fill=tk.X, pady=5)
        
        tk.Label(cf, text="", width=6, bg=self.BG_COLOR).grid(row=0, column=0)
        for j in range(nd):
            tk.Label(cf, text="D" + str(j+1), font=('Segoe UI', 10, 'bold'),
                    width=8, bg=self.TABLE_HEADER_COLOR, relief=tk.RIDGE).grid(row=0, column=j+1, padx=2, pady=2, sticky="nsew")
        
        sample = [[3, 7, 2, 5], [4, 1, 8, 6], [9, 3, 4, 2]]
        for i in range(ns):
            tk.Label(cf, text="S" + str(i+1), font=('Segoe UI', 10, 'bold'),
                    width=6, bg=self.TABLE_HEADER_COLOR, relief=tk.RIDGE).grid(row=i+1, column=0, padx=2, pady=2, sticky="nsew")
            for j in range(nd):
                e = tk.Entry(cf, width=8, font=('Segoe UI', 11), justify=tk.CENTER, relief=tk.SOLID, bd=1)
                e.grid(row=i+1, column=j+1, padx=2, pady=2)
                if i < len(sample) and j < len(sample[i]):
                    e.insert(0, str(sample[i][j]))
                else:
                    e.insert(0, "0")
                self.cost_entries[i][j] = e
        
        # Buttons
        bf = tk.Frame(scroll_frame, bg=self.BG_COLOR)
        bf.pack(pady=20)
        tk.Button(bf, text="<- Back to Setup", font=('Segoe UI', 10),
                 bg='#757575', fg='white', padx=15, pady=5,
                 command=lambda: self.notebook.select(self.page1)).pack(side=tk.LEFT, padx=10)
        tk.Button(bf, text="NEXT  ->  Solve Problem", font=('Segoe UI', 12, 'bold'),
                 bg=self.BUTTON_COLOR, fg='white', padx=25, pady=8,
                 command=self._read_data_and_go).pack(side=tk.LEFT, padx=10)
    
    def _read_data_and_go(self):
        try:
            supply = [int(e.get()) for e in self.supply_entries]
            demand = [int(e.get()) for e in self.demand_entries]
            costs = []
            for i in range(self.num_supply.get()):
                row = []
                for j in range(self.num_demand.get()):
                    row.append(int(self.cost_entries[i][j].get()))
                costs.append(row)
            for s in supply:
                if s <= 0: raise ValueError("Supply must be positive")
            for d in demand:
                if d <= 0: raise ValueError("Demand must be positive")
            for row in costs:
                for c in row:
                    if c < 0: raise ValueError("Cost cannot be negative")
            self.entered_supply = supply
            self.entered_demand = demand
            self.entered_costs = costs
            self._create_page3_solve()
            self.notebook.select(self.page3)
        except ValueError as e:
            messagebox.showerror("Input Error", str(e))
        except Exception as e:
            messagebox.showerror("Error", "Invalid input: " + str(e))
    
    # ============================
    # PAGE 3
    # ============================
    
    def _create_page3_solve(self):
        for widget in self.page3.winfo_children():
            widget.destroy()
        
        frame = tk.Frame(self.page3, bg=self.BG_COLOR)
        frame.pack(expand=True, pady=40)
        
        tk.Label(frame, text="Solve the Problem", font=('Segoe UI', 18, 'bold'),
                bg=self.BG_COLOR, fg=self.HEADER_COLOR).pack(pady=(0, 30))
        
        sframe = tk.Frame(frame, bg='#e3f2fd', bd=1, relief=tk.SOLID)
        sframe.pack(pady=10, ipadx=30, ipady=10)
        
        total_str = "Origins: " + str(len(self.entered_supply)) + "  |  Destinations: " + str(len(self.entered_demand))
        total_str += "  |  Total Supply: " + str(sum(self.entered_supply))
        total_str += "  |  Total Demand: " + str(sum(self.entered_demand))
        
        tk.Label(sframe, text=total_str, font=('Segoe UI', 11), bg='#e3f2fd', fg='#1565c0').pack()
        
        if sum(self.entered_supply) == sum(self.entered_demand):
            tk.Label(sframe, text="OK - Problem is BALANCED", 
                    font=('Segoe UI', 10, 'bold'), bg='#e3f2fd', fg='#2e7d32').pack()
        else:
            tk.Label(sframe, text="NOTE: Problem is UNBALANCED (will add dummy row/column)", 
                    font=('Segoe UI', 10, 'bold'), bg='#e3f2fd', fg='#e65100').pack()
        
        tk.Label(frame, text="Select Methods:", font=('Segoe UI', 14, 'bold'),
                bg=self.BG_COLOR, fg=self.HEADER_COLOR).pack(pady=(30, 15))
        
        self.method_vars = {
            'nwc': tk.BooleanVar(value=True),
            'lcm': tk.BooleanVar(value=True),
            'vam': tk.BooleanVar(value=True),
            'modi': tk.BooleanVar(value=True)
        }
        
        mf = tk.Frame(frame, bg=self.BG_COLOR)
        mf.pack(pady=10)
        
        methods = [
            ('nwc', 'NWC (North-West Corner)', 'Simple'),
            ('lcm', 'LCM (Least Cost)', 'Medium'),
            ('vam', 'VAM (Vogel)', 'Best initial  *'),
            ('modi', 'MODI (Modified Distribution)', 'Optimal  **')
        ]
        
        for key, name, desc in methods:
            row = tk.Frame(mf, bg=self.BG_COLOR)
            row.pack(fill=tk.X, pady=4)
            cb = tk.Checkbutton(row, text=name, variable=self.method_vars[key],
                              font=('Segoe UI', 11), bg=self.BG_COLOR, anchor=tk.W, width=40)
            cb.pack(side=tk.LEFT)
            tk.Label(row, text=desc, font=('Segoe UI', 10), fg='#757575', bg=self.BG_COLOR).pack(side=tk.LEFT, padx=10)
        
        tk.Button(frame, text=">>  SOLVE", font=('Segoe UI', 14, 'bold'),
                 bg='#2e7d32', fg='white', padx=40, pady=12,
                 command=self._solve_problem).pack(pady=30)
        
        tk.Button(frame, text="<- Back to Data Entry", font=('Segoe UI', 10),
                 bg='#757575', fg='white', padx=15, pady=5,
                 command=lambda: (self._create_page2_data(), self.notebook.select(self.page2))).pack()
    
    def _solve_problem(self):
        supply = self.entered_supply
        demand = self.entered_demand
        costs = self.entered_costs
        self.results = {}
        
        try:
            if self.method_vars['nwc'].get():
                s = TransportSolver(supply, demand, costs)
                sol, cost, steps = s.nwc()
                self.results['NWC'] = {'solver': s, 'solution': sol, 'cost': cost, 'steps': steps}
            if self.method_vars['lcm'].get():
                s = TransportSolver(supply, demand, costs)
                sol, cost, steps = s.least_cost()
                self.results['LCM'] = {'solver': s, 'solution': sol, 'cost': cost, 'steps': steps}
            if self.method_vars['vam'].get():
                s = TransportSolver(supply, demand, costs)
                sol, cost, steps = s.vogel()
                self.results['VAM'] = {'solver': s, 'solution': sol, 'cost': cost, 'steps': steps}
            if self.method_vars['modi'].get():
                s = TransportSolver(supply, demand, costs)
                sol, cost, steps = s.modi('vogel')
                self.results['MODI'] = {'solver': s, 'solution': sol, 'cost': cost, 'steps': steps}
            
            if not self.results:
                messagebox.showwarning("No Method", "Please select at least one method.")
                return
            
            self._create_page4_results()
            self.notebook.select(self.page4)
        except Exception as e:
            messagebox.showerror("Solver Error", "An error occurred: " + str(e))
            import traceback
            traceback.print_exc()
    
    # ============================
    # PAGE 4
    # ============================
    
    def _create_page4_results(self):
        for widget in self.page4.winfo_children():
            widget.destroy()
        
        main_frame = tk.Frame(self.page4, bg=self.BG_COLOR)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        tk.Label(main_frame, text="Results", font=('Segoe UI', 16, 'bold'),
                bg=self.BG_COLOR, fg=self.HEADER_COLOR).pack(anchor=tk.W, pady=(0, 10))
        
        if not self.results:
            tk.Label(main_frame, text="No results. Go back and solve first.",
                    font=('Segoe UI', 12), bg=self.BG_COLOR, fg='red').pack()
            return
        
        best_method = min(self.results, key=lambda k: self.results[k]['cost'])
        best_cost = self.results[best_method]['cost']
        
        sframe = tk.Frame(main_frame, bg='#e8f5e9', bd=1, relief=tk.SOLID)
        sframe.pack(fill=tk.X, pady=5, ipady=5)
        tk.Label(sframe, text="* BEST: " + best_method + " -> Total Cost = " + str(best_cost),
                font=('Segoe UI', 13, 'bold'), bg='#e8f5e9', fg='#2e7d32').pack(pady=5)
        
        rn = ttk.Notebook(main_frame)
        rn.pack(fill=tk.BOTH, expand=True, pady=10)
        
        for method_name in ['NWC', 'LCM', 'VAM', 'MODI']:
            if method_name not in self.results:
                continue
            result = self.results[method_name]
            page = tk.Frame(rn, bg='white')
            rn.add(page, text=" " + method_name + "  ")
            self._fill_result_page(page, method_name, result)
        
        bf = tk.Frame(main_frame, bg=self.BG_COLOR)
        bf.pack(pady=10)
        tk.Button(bf, text="<- Back to Solve", font=('Segoe UI', 10),
                 bg='#757575', fg='white', padx=15, pady=5,
                 command=lambda: self.notebook.select(self.page3)).pack(side=tk.LEFT, padx=10)
        tk.Button(bf, text="Next -> Compare All", font=('Segoe UI', 11, 'bold'),
                 bg=self.BUTTON_COLOR, fg='white', padx=20, pady=7,
                 command=lambda: (self._create_page5_compare(), self.notebook.select(self.page5))).pack(side=tk.LEFT, padx=10)
    
    def _fill_result_page(self, parent, method_name, result):
        solver = result['solver']
        cost = result['cost']
        solution = result['solution']
        ns = solver.num_supply
        nd = solver.num_demand
        
        cframe = tk.Frame(parent, bg='white')
        cframe.pack(fill=tk.X, pady=10, padx=10)
        tk.Label(cframe, text="Total Cost: " + str(cost), font=('Segoe UI', 14, 'bold'),
                fg='#2e7d32', bg='white').pack(side=tk.LEFT, padx=10)
        tk.Label(cframe, text="Method: " + solver.method_used, font=('Segoe UI', 11),
                fg='#757575', bg='white').pack(side=tk.RIGHT, padx=10)
        
        tframe = tk.Frame(parent, bg='white')
        tframe.pack(pady=10, padx=10)
        
        table = tk.Frame(tframe, bg='#e0e0e0', bd=1, relief=tk.SOLID)
        table.pack()
        
        tk.Label(table, text="", width=8, bg=self.TABLE_HEADER_COLOR, 
                font=('Segoe UI', 9, 'bold'), relief=tk.RIDGE).grid(row=0, column=0)
        for j in range(nd):
            tk.Label(table, text="D" + str(j+1), width=8, bg=self.TABLE_HEADER_COLOR,
                    font=('Segoe UI', 9, 'bold'), relief=tk.RIDGE).grid(row=0, column=j+1)
        tk.Label(table, text="Supply", width=8, bg=self.TABLE_HEADER_COLOR,
                font=('Segoe UI', 9, 'bold'), relief=tk.RIDGE).grid(row=0, column=nd+1)
        
        for i in range(ns):
            bg = 'white' if i % 2 == 0 else self.ROW_ALT_COLOR
            tk.Label(table, text="S" + str(i+1), width=8, bg=self.TABLE_HEADER_COLOR,
                    font=('Segoe UI', 9, 'bold'), relief=tk.RIDGE).grid(row=i+1, column=0)
            for j in range(nd):
                val = solution[i][j]
                fg_c = '#2e7d32' if val > 0 else '#e0e0e0'
                txt = str(int(val)) if val == int(val) else "{:.1f}".format(val)
                tk.Label(table, text=txt, width=8, bg=bg, fg=fg_c,
                        font=('Segoe UI', 10, 'bold' if val > 0 else 'normal'),
                        relief=tk.RIDGE).grid(row=i+1, column=j+1)
            tk.Label(table, text=str(solver.supply[i]), width=8, bg=bg,
                    font=('Segoe UI', 9, 'bold'), relief=tk.RIDGE).grid(row=i+1, column=nd+1)
        
        tk.Label(table, text="Demand", width=8, bg=self.TABLE_HEADER_COLOR,
                font=('Segoe UI', 9, 'bold'), relief=tk.RIDGE).grid(row=ns+1, column=0)
        for j in range(nd):
            tk.Label(table, text=str(solver.demand[j]), width=8, bg=self.TABLE_HEADER_COLOR,
                    font=('Segoe UI', 9, 'bold'), relief=tk.RIDGE).grid(row=ns+1, column=j+1)
        tk.Label(table, text="", width=8, bg=self.TABLE_HEADER_COLOR, relief=tk.RIDGE).grid(row=ns+1, column=nd+1)
        
        aframe = tk.Frame(parent, bg='white')
        aframe.pack(fill=tk.X, padx=10, pady=10)
        tk.Label(aframe, text="Allocations:", font=('Segoe UI', 11, 'bold'), bg='white').pack(anchor=tk.W)
        
        lf = tk.Frame(aframe, bg='white', bd=1, relief=tk.SUNKEN)
        lf.pack(fill=tk.X, pady=5)
        
        canvas = tk.Canvas(lf, bg='white', height=100, highlightthickness=0)
        sb = tk.Scrollbar(lf, orient=tk.VERTICAL, command=canvas.yview)
        sc = tk.Frame(canvas, bg='white')
        sc.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=sc, anchor=tk.NW)
        canvas.configure(yscrollcommand=sb.set)
        
        summary = solver.get_solution_summary()
        for idx2, assign in enumerate(summary['assignments']):
            bg2 = 'white' if idx2 % 2 == 0 else self.ROW_ALT_COLOR
            row_f = tk.Frame(sc, bg=bg2)
            row_f.pack(fill=tk.X)
            txt2 = assign['from'] + " -> " + assign['to'] + ":  " + str(assign['amount']) + " units x $" + str(assign['unit_cost']) + " = $" + str(assign['total_cost'])
            tk.Label(row_f, text=txt2, font=('Segoe UI', 10), bg=bg2, anchor=tk.W).pack(padx=10, pady=2)
        
        canvas.pack(side=tk.LEFT, fill=tk.X, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
    
   
        # ============================
    # PAGE 5
    # ============================
    
    def _create_page5_compare(self):
        for widget in self.page5.winfo_children():
            widget.destroy()
        
        frame = tk.Frame(self.page5, bg=self.BG_COLOR)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tk.Label(frame, text="Compare & Export", font=('Segoe UI', 18, 'bold'),
                bg=self.BG_COLOR, fg=self.HEADER_COLOR).pack(anchor=tk.W, pady=(0, 20))
        
        if not self.results:
            tk.Label(frame, text="No results to compare. Solve first.", 
                    font=('Segoe UI', 12), bg=self.BG_COLOR).pack()
            return
        
        # --- Problem Summary ---
        sframe = tk.Frame(frame, bg='#e3f2fd', bd=1, relief=tk.SOLID)
        sframe.pack(fill=tk.X, pady=5, ipady=5)
        
        total_str = "Origins: " + str(len(self.entered_supply)) + "  |  Destinations: " + str(len(self.entered_demand))
        total_str += "  |  Total Supply: " + str(sum(self.entered_supply))
        total_str += "  |  Total Demand: " + str(sum(self.entered_demand))
        tk.Label(sframe, text=total_str, font=('Segoe UI', 11), bg='#e3f2fd', fg='#1565c0').pack()
        
        # --- Comparison Table ---
        cf = tk.Frame(frame, bg='white', bd=1, relief=tk.SOLID)
        cf.pack(pady=15)
        
        headers = ["Method", "Total Cost", "Allocations", "Balanced", "Status"]
        widths = [25, 14, 12, 10, 12]
        
        for col, (hdr, w) in enumerate(zip(headers, widths)):
            tk.Label(cf, text=hdr, width=w, bg=self.TABLE_HEADER_COLOR,
                    font=('Segoe UI', 10, 'bold'), relief=tk.RIDGE).grid(row=0, column=col, sticky="nsew")
        
        best_cost = min(self.results[k]['cost'] for k in self.results)
        row_idx = 1
        
        for method_name in ['NWC', 'LCM', 'VAM', 'MODI']:
            if method_name not in self.results:
                continue
            result = self.results[method_name]
            cost = result['cost']
            solver = result['solver']
            is_best = (cost == best_cost)
            bg = 'white' if row_idx % 2 == 0 else self.ROW_ALT_COLOR
            
            alloc_count = sum(1 for row in result['solution'] for v in row if v > 0)
            
            tk.Label(cf, text=method_name, width=widths[0], bg=bg,
                    font=('Segoe UI', 10, 'bold' if is_best else 'normal'),
                    fg=self.HEADER_COLOR if is_best else 'black', relief=tk.RIDGE).grid(row=row_idx, column=0, sticky="nsew")
            
            tk.Label(cf, text="$ " + str(cost), width=widths[1], bg=bg,
                    font=('Segoe UI', 10, 'bold' if is_best else 'normal'),
                    fg='#2e7d32' if is_best else 'black', relief=tk.RIDGE).grid(row=row_idx, column=1)
            
            tk.Label(cf, text=str(alloc_count), width=widths[2], bg=bg,
                    font=('Segoe UI', 10), relief=tk.RIDGE).grid(row=row_idx, column=2)
            
            tk.Label(cf, text="YES" if solver.is_balanced_flag else "NO", 
                    width=widths[3], bg=bg,
                    font=('Segoe UI', 10), relief=tk.RIDGE).grid(row=row_idx, column=3)
            
            if method_name == "MODI":
                status = "OPTIMAL *" if is_best else "CHECK"
            else:
                status = "BEST INITIAL" if is_best else ""
            tk.Label(cf, text=status, width=widths[4], bg=bg,
                    font=('Segoe UI', 10, 'bold'), fg='#2e7d32' if is_best else '#757575',
                    relief=tk.RIDGE).grid(row=row_idx, column=4)
            
            row_idx += 1
        
        # --- Best method ---
        best_frame = tk.Frame(frame, bg='#e8f5e9', bd=1, relief=tk.SOLID)
        best_frame.pack(fill=tk.X, pady=8, ipady=5)
        best_method = min(self.results, key=lambda k: self.results[k]['cost'])
        tk.Label(best_frame, 
                text="* BEST FOUND: " + best_method + "  |  Total Cost = $ " + str(self.results[best_method]['cost']),
                font=('Segoe UI', 12, 'bold'), bg='#e8f5e9', fg='#2e7d32').pack(pady=3)
        
        # --- Export Buttons ---
        btn_frame = tk.Frame(frame, bg=self.BG_COLOR)
        btn_frame.pack(pady=15)
        
        tk.Button(btn_frame, text="Export TXT", font=('Segoe UI', 11, 'bold'),
                 bg='#1565c0', fg='white', padx=20, pady=7,
                 command=self._export_txt).pack(side=tk.LEFT, padx=8)
        
        tk.Button(btn_frame, text="Export PDF", font=('Segoe UI', 11, 'bold'),
                 bg='#c62828', fg='white', padx=20, pady=7,
                 command=self._export_pdf).pack(side=tk.LEFT, padx=8)
        
        tk.Button(btn_frame, text="Export Excel", font=('Segoe UI', 11, 'bold'),
                 bg='#e65100', fg='white', padx=20, pady=7,
                 command=self._export_excel).pack(side=tk.LEFT, padx=8)
        
        # --- Navigation ---
        nav_frame = tk.Frame(frame, bg=self.BG_COLOR)
        nav_frame.pack(pady=10)
        tk.Button(nav_frame, text="<- Back to Results", font=('Segoe UI', 10),
                 bg='#757575', fg='white', padx=15, pady=5,
                 command=lambda: self.notebook.select(self.page4)).pack(side=tk.LEFT, padx=10)
    
    def _export_txt(self):
        """Export results to a text file"""
        if not self.results:
            messagebox.showwarning("No Data", "Nothing to export.")
            return
        
        try:
            from datetime import datetime
            lines = []
            lines.append("=" * 60)
            lines.append("TRANSPORTATION PROBLEM - RESULTS")
            lines.append("=" * 60)
            lines.append("")
            lines.append("Prepared by Navid Naderpour,navid.naderpour@gmail.com")
            lines.append("Date: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            lines.append("")
            
            supply = self.entered_supply
            demand = self.entered_demand
            costs = self.entered_costs
            
            lines.append("Supply: " + str(supply))
            lines.append("Demand: " + str(demand))
            lines.append("Cost Matrix:")
            for i, row in enumerate(costs):
                lines.append("  S" + str(i+1) + ": " + str(row))
            lines.append("")
            
            for method_name in ['NWC', 'LCM', 'VAM', 'MODI']:
                if method_name not in self.results:
                    continue
                result = self.results[method_name]
                solver = result['solver']
                lines.append("-" * 50)
                lines.append("Method: " + solver.method_used)
                lines.append("Total Cost: " + str(result['cost']))
                lines.append("")
                
                # Solution matrix
                sol = result['solution']
                ns2 = solver.num_supply
                nd2 = solver.num_demand
                lines.append("Solution Matrix:")
                header = " " * 10
                for j in range(nd2):
                    header += "D" + str(j+1) + "  "
                header += "Supply"
                lines.append(header)
                for i in range(ns2):
                    row_txt = "S" + str(i+1) + " " * 6
                    for j in range(nd2):
                        if sol[i][j] > 0:
                            row_txt += str(int(sol[i][j])) + "   "
                        else:
                            row_txt += ".   "
                    row_txt += str(solver.supply[i])
                    lines.append(row_txt)
                
                # Allocations
                summary = solver.get_solution_summary()
                lines.append("")
                lines.append("Allocations:")
                for assign in summary['assignments']:
                    lines.append("  " + assign['from'] + " -> " + assign['to'] + ": " + str(assign['amount']) + " units")
                lines.append("")
            
                # Best method
            best_method = min(self.results, key=lambda k: self.results[k]['cost'])
            lines.append("=" * 60)
            lines.append("BEST METHOD: " + best_method + " with Total Cost = " + str(self.results[best_method]['cost']))
            lines.append("=" * 60)
            
            text = "\n".join(lines)
            
            # Save to Desktop\Navid_OR
            import os
            # Save in current folder (where the script is running)
            desktop = os.getcwd()
            if not os.path.exists(desktop):
                os.makedirs(desktop)
            from tkinter import filedialog
            default_name = "transport_results_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".txt"
            filename = filedialog.asksaveasfilename(
                initialdir=desktop,
                initialfile=default_name,
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                title="Save TXT Report"
            )
            if not filename:
                return  # User cancelled
            
            with open(filename, "w", encoding="utf-8") as f:
                f.write(text)
            
            messagebox.showinfo("Exported", "Results saved to:\n" + filename)
            
        except Exception as e:
            messagebox.showerror("Export Error", "Could not export: " + str(e))
    def _export_pdf(self):
        """Export a professional PDF report with full results and analysis"""
        if not self.results:
            messagebox.showwarning("No Data", "Nothing to export. Solve first.")
            return
        
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, mm
            from reportlab.lib.colors import HexColor, black, white
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, 
                                           TableStyle, PageBreak, HRFlowable)
            from reportlab.lib import colors
            from datetime import datetime
            import os
            
            # Save in current folder (where the script is running)
            desktop = os.getcwd()
            if not os.path.exists(desktop):
                os.makedirs(desktop)
            
            from tkinter import filedialog
            default_name = "transport_report_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".pdf"
            filename = filedialog.asksaveasfilename(
                initialdir=desktop,
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                title="Save PDF Report"
            )
            if not filename:
                return  # User cancelled
            doc = SimpleDocTemplate(filename, pagesize=A4,
                                   leftMargin=40, rightMargin=40,
                                   topMargin=40, bottomMargin=40)
            
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                        fontSize=20, textColor=HexColor('#1a237e'),
                                        spaceAfter=12, spaceBefore=6)
            heading1_style = ParagraphStyle('H1', parent=styles['Heading1'],
                                           fontSize=16, textColor=HexColor('#1a237e'),
                                           spaceAfter=10, spaceBefore=16)
            heading2_style = ParagraphStyle('H2', parent=styles['Heading2'],
                                           fontSize=13, textColor=HexColor('#283593'),
                                           spaceAfter=8, spaceBefore=12)
            heading3_style = ParagraphStyle('H3', parent=styles['Heading3'],
                                           fontSize=11, textColor=HexColor('#1565c0'),
                                           spaceAfter=6, spaceBefore=10)
            normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'],
                                         fontSize=10, spaceAfter=4, leading=14)
            bold_style = ParagraphStyle('BoldNormal', parent=normal_style,
                                       fontName='Helvetica-Bold')
            analysis_style = ParagraphStyle('Analysis', parent=normal_style,
                                           backColor=HexColor('#f3e5f5'),
                                           borderPadding=10, borderWidth=1,
                                           borderColor=HexColor('#ce93d8'),
                                           spaceAfter=10, spaceBefore=10)
            summary_style = ParagraphStyle('Summary', parent=normal_style,
                                          backColor=HexColor('#e8f5e9'),
                                          borderPadding=10, borderWidth=1,
                                          borderColor=HexColor('#a5d6a7'),
                                          spaceAfter=10, spaceBefore=10)
            footer_style = ParagraphStyle('Footer', parent=styles['Normal'],
                                         fontSize=8, textColor=HexColor('#757575'),
                                         alignment=1, spaceBefore=20)
            
            elements = []
            
            # Title
            elements.append(Paragraph("Transportation Problem Report", title_style))
            elements.append(Paragraph("<i>Prepared by Navid Naderpour,navid.naderpour@gmail.com</i>", normal_style))
            elements.append(Paragraph("<i>Date: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "</i>", normal_style))
            elements.append(Spacer(1, 12))
            elements.append(HRFlowable(width="100%", thickness=2, color=HexColor('#1a237e')))
            elements.append(Spacer(1, 12))
            
            # 1. Problem Data
            elements.append(Paragraph("1. Problem Data", heading1_style))
            
            supply_data = [["Origin", "Supply"]]
            for i, s in enumerate(self.entered_supply):
                supply_data.append(["S" + str(i+1), str(s)])
            supply_table = Table(supply_data, colWidths=[100, 100])
            supply_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#e8eaf6')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(Paragraph("Supply (Origins):", heading3_style))
            elements.append(supply_table)
            elements.append(Spacer(1, 8))
            
            demand_data = [["Destination", "Demand"]]
            for j, d in enumerate(self.entered_demand):
                demand_data.append(["D" + str(j+1), str(d)])
            demand_table = Table(demand_data, colWidths=[100, 100])
            demand_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#e8eaf6')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(Paragraph("Demand (Destinations):", heading3_style))
            elements.append(demand_table)
            elements.append(Spacer(1, 8))
            
            total_supply = sum(self.entered_supply)
            total_demand = sum(self.entered_demand)
            if total_supply == total_demand:
                elements.append(Paragraph("<b>Status:</b> Balanced (Total = " + str(total_supply) + ")", normal_style))
            else:
                elements.append(Paragraph("<b>Status:</b> Unbalanced. A dummy row/column was added.", normal_style))
            elements.append(Spacer(1, 8))
            
            # Cost matrix
            cost_header = [""] + ["D" + str(j+1) for j in range(len(self.entered_demand))]
            cost_data = [cost_header]
            for i, row in enumerate(self.entered_costs):
                cost_data.append(["S" + str(i+1)] + [str(c) for c in row])
            
            cost_table = Table(cost_data)
            cost_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#e8eaf6')),
                ('BACKGROUND', (0, 1), (0, -1), HexColor('#e8eaf6')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(Paragraph("Cost Matrix:", heading3_style))
            elements.append(cost_table)
            elements.append(Spacer(1, 12))
            
            # 2. Results by Method
            elements.append(Paragraph("2. Results by Method", heading1_style))
            
            best_cost = min(self.results[k]['cost'] for k in self.results)
            best_method = min(self.results, key=lambda k: self.results[k]['cost'])
            
            for method_name in ['NWC', 'LCM', 'VAM', 'MODI']:
                if method_name not in self.results:
                    continue
                result = self.results[method_name]
                solver = result['solver']
                sol = result['solution']
                cost = result['cost']
                is_best = (cost == best_cost)
                
                title = solver.method_used + " — Total Cost: $" + str(cost)
                if is_best:
                    title += " (OPTIMAL)"
                elements.append(Paragraph(title, heading2_style))
                
                # Solution matrix
                sol_header = [""] + ["D" + str(j+1) for j in range(solver.num_demand)] + ["Supply"]
                sol_data = [sol_header]
                for i in range(solver.num_supply):
                    row = ["S" + str(i+1)]
                    for j in range(solver.num_demand):
                        val = sol[i][j]
                        if val > 0:
                            display = str(int(val)) if val == int(val) else "{:.1f}".format(val)
                            row.append(display)
                        else:
                            row.append("—")
                    row.append(str(solver.supply[i]))
                    sol_data.append(row)
                
                demand_row = ["Demand"]
                for j in range(solver.num_demand):
                    demand_row.append(str(solver.demand[j]))
                demand_row.append("")
                sol_data.append(demand_row)
                
                sol_table = Table(sol_data)
                sol_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), HexColor('#e8eaf6')),
                    ('BACKGROUND', (0, 1), (0, -1), HexColor('#e8eaf6')),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(sol_table)
                elements.append(Spacer(1, 6))
                
                # Allocations
                elements.append(Paragraph("Allocations:", bold_style))
                summary = solver.get_solution_summary()
                for assign in summary['assignments']:
                    a = assign['amount']
                    a_str = str(int(a)) if a == int(a) else "{:.1f}".format(a)
                    elements.append(Paragraph(
                        "&nbsp;&nbsp;&nbsp;" + assign['from'] + " → " + assign['to'] + ": " + a_str + " units × $" + str(assign['unit_cost']) + " = $" + str(assign['total_cost']),
                        normal_style))
                
                elements.append(Spacer(1, 10))
            
            # 3. Comparison
            elements.append(HRFlowable(width="100%", thickness=1, color=HexColor('#c5cae9')))
            elements.append(Spacer(1, 6))
            elements.append(Paragraph("3. Method Comparison", heading1_style))
            
            comp_header = ["Method", "Total Cost", "Allocations", "Status"]
            comp_data = [comp_header]
            for method_name in ['NWC', 'LCM', 'VAM', 'MODI']:
                if method_name not in self.results:
                    continue
                cost = self.results[method_name]['cost']
                is_best = (cost == best_cost)
                alloc_count = sum(1 for row in self.results[method_name]['solution'] for v in row if v > 0)
                
                comp_data.append([method_name, "$" + str(cost), str(alloc_count), "OPTIMAL" if is_best else "—"])
            
            comp_table = Table(comp_data, colWidths=[120, 100, 100, 100])
            comp_style = [
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#e8eaf6')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]
            # Highlight best row
            for idx in range(1, len(comp_data)):
                if comp_data[idx][3] == "OPTIMAL":
                    comp_style.append(('BACKGROUND', (0, idx), (-1, idx), HexColor('#e8f5e9')))
                    comp_style.append(('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'))
            
            comp_table.setStyle(TableStyle(comp_style))
            elements.append(comp_table)
            elements.append(Spacer(1, 10))
            
            # Summary
            elements.append(Paragraph("Summary", heading1_style))
            elements.append(Paragraph("<b>Optimal Method:</b> " + best_method, summary_style))
            elements.append(Paragraph("<b>Minimum Total Cost:</b> $" + str(self.results[best_method]['cost']), summary_style))
            elements.append(Paragraph("<b>Methods Compared:</b> " + str(len(self.results)), summary_style))
            
            elements.append(Spacer(1, 10))
            
            # 4. Analysis
            elements.append(Paragraph("4. Analysis", heading1_style))
            
            costs_dict = {m: self.results[m]['cost'] for m in self.results}
            sorted_methods = sorted(costs_dict, key=lambda m: costs_dict[m])
            
            analysis_lines = []
            analysis_lines.append("<b>Cost Efficiency Analysis:</b><br/><br/>")
            
            for idx, m in enumerate(sorted_methods):
                c = costs_dict[m]
                if idx == 0:
                    analysis_lines.append("• <b>" + m + "</b> achieved the optimal cost of <b>$" + str(c) + "</b>.<br/>")
                else:
                    pct = round((c - costs_dict[sorted_methods[0]]) / costs_dict[sorted_methods[0]] * 100, 1)
                    analysis_lines.append("• <b>" + m + "</b> resulted in <b>$" + str(c) + "</b> (" + str(pct) + "% higher).<br/>")
            
            first = sorted_methods[0]
            last = sorted_methods[-1]
            savings = costs_dict[last] - costs_dict[first]
            savings_pct = round((costs_dict[last] - costs_dict[first]) / costs_dict[last] * 100, 1)
            
            analysis_lines.append("<br/><b>Key Findings:</b><br/><br/>")
            analysis_lines.append("• Using <b>" + first + "</b> saves <b>$" + str(savings) + " (" + str(savings_pct) + "%)</b> over <b>" + last + "</b>.<br/>")
            
            if 'MODI' in costs_dict and 'VAM' in costs_dict:
                if costs_dict['MODI'] == costs_dict['VAM']:
                    analysis_lines.append("• MODI confirmed VAM was already optimal.<br/>")
                else:
                    analysis_lines.append("• MODI improved VAM from $" + str(costs_dict['VAM']) + " to $" + str(costs_dict['MODI']) + ".<br/>")
            
            if total_supply != total_demand:
                analysis_lines.append("• A dummy variable was added due to imbalance.<br/>")
            
            analysis_lines.append("<br/><b>Recommendation:</b> Use <b>" + first + "</b> to minimize total cost to <b>$" + str(costs_dict[first]) + "</b>.")
            
            elements.append(Paragraph("".join(analysis_lines), analysis_style))
            
            # Footer
            elements.append(Spacer(1, 20))
            elements.append(HRFlowable(width="100%", thickness=1, color=HexColor('#e0e0e0')))
            elements.append(Paragraph("Transportation Problem Solver — Prepared by Navid Naderpour,navid.naderpour@gmail.com", footer_style))
            elements.append(Paragraph("Generated on " + datetime.now().strftime("%Y-%m-%d at %H:%M:%S"), footer_style))
            
            doc.build(elements)
            
            messagebox.showinfo("PDF Generated", "PDF report saved to:\n" + filename)
            
        except Exception as e:
            messagebox.showerror("PDF Error", "Could not generate PDF:\n" + str(e) + 
                               "\n\nMake sure 'reportlab' is installed:\npip install reportlab")
    
    def _export_excel(self):
        """Export results to an Excel file with full data and analysis"""
        if not self.results:
            messagebox.showwarning("No Data", "Nothing to export. Solve first.")
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            import os

            desktop = os.getcwd()
            wb = Workbook()
            
            header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
            header_font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
            sub_header_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
            sub_header_font = Font(name='Segoe UI', bold=True, color='1A237E', size=10)
            normal_font = Font(name='Segoe UI', size=10)
            bold_font = Font(name='Segoe UI', bold=True, size=10)
            green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            green_font = Font(name='Segoe UI', bold=True, color='2E7D32', size=10)
            analysis_fill = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
            summary_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin', color='C5CAE9'),
                right=Side(style='thin', color='C5CAE9'),
                top=Side(style='thin', color='C5CAE9'),
                bottom=Side(style='thin', color='C5CAE9')
            )
            
            def style_header_row(ws, row, max_col):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
            
            def style_sub_header_row(ws, row, max_col):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = sub_header_fill
                    cell.font = sub_header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
            
            def style_data_cell(ws, row, col, bold=False, fill=None):
                cell = ws.cell(row=row, column=col)
                cell.font = bold_font if bold else normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                if fill:
                    cell.fill = fill
            
            ws1 = wb.active
            ws1.title = "Report"
            
            ws1.column_dimensions['A'].width = 25
            ws1.column_dimensions['B'].width = 18
            ws1.column_dimensions['C'].width = 18
            ws1.column_dimensions['D'].width = 18
            ws1.column_dimensions['E'].width = 18
            ws1.column_dimensions['F'].width = 18
            
            row = 1
            
            # Title
            ws1.merge_cells('A1:F1')
            ws1.cell(row=1, column=1, value="Transportation Problem Report").font = Font(name='Segoe UI', bold=True, color='1A237E', size=16)
            ws1.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            
            row = 3
            ws1.cell(row=row, column=1, value="Prepared by: Navid Naderpour,navid.naderpour@gmail.com").font = Font(name='Segoe UI', size=10, italic=True)
            row = 4
            ws1.cell(row=row, column=1, value="Date: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font = Font(name='Segoe UI', size=10, italic=True)
            
            # 1. Problem Data
            row = 6
            ws1.cell(row=row, column=1, value="1. Problem Data").font = Font(name='Segoe UI', bold=True, color='1A237E', size=13)
            
            row = 8
            ws1.cell(row=row, column=1, value="Supply (Origins):")
            style_sub_header_row(ws1, row, 1)
            row = 9
            ws1.cell(row=row, column=1, value="Origin")
            ws1.cell(row=row, column=2, value="Supply")
            style_sub_header_row(ws1, row, 2)
            
            for i, s in enumerate(self.entered_supply):
                row = 10 + i
                ws1.cell(row=row, column=1, value="S" + str(i+1))
                ws1.cell(row=row, column=2, value=s)
                style_data_cell(ws1, row, 1)
                style_data_cell(ws1, row, 2)
            
            row = 10 + len(self.entered_supply)
            ws1.cell(row=row, column=1, value="Demand (Destinations):")
            style_sub_header_row(ws1, row, 1)
            row += 1
            ws1.cell(row=row, column=1, value="Destination")
            ws1.cell(row=row, column=2, value="Demand")
            style_sub_header_row(ws1, row, 2)
            
            for j, d in enumerate(self.entered_demand):
                row += 1
                ws1.cell(row=row, column=1, value="D" + str(j+1))
                ws1.cell(row=row, column=2, value=d)
                style_data_cell(ws1, row, 1)
                style_data_cell(ws1, row, 2)
            
            total_supply = sum(self.entered_supply)
            total_demand = sum(self.entered_demand)
            
            row += 2
            if total_supply == total_demand:
                ws1.cell(row=row, column=1, value="Status: Balanced (Total = " + str(total_supply) + ")").font = green_font
            else:
                ws1.cell(row=row, column=1, value="Status: Unbalanced (dummy added)").font = Font(name='Segoe UI', bold=True, color='E65100', size=10)
            
            # Cost matrix
            row += 2
            ws1.cell(row=row, column=1, value="Cost Matrix:")
            style_sub_header_row(ws1, row, 1)
            row += 1
            ws1.cell(row=row, column=1, value="")
            for j in range(len(self.entered_demand)):
                ws1.cell(row=row, column=j+2, value="D" + str(j+1))
            style_sub_header_row(ws1, row, len(self.entered_demand) + 1)
            
            for i, cost_row in enumerate(self.entered_costs):
                row += 1
                ws1.cell(row=row, column=1, value="S" + str(i+1))
                style_data_cell(ws1, row, 1, bold=True)
                for j, c in enumerate(cost_row):
                    ws1.cell(row=row, column=j+2, value=c)
                    style_data_cell(ws1, row, j+2)
            
            # 2. Results by Method
            row += 3
            ws1.cell(row=row, column=1, value="2. Results by Method").font = Font(name='Segoe UI', bold=True, color='1A237E', size=13)
            
            best_cost = min(self.results[k]['cost'] for k in self.results)
            
            for method_name in ['NWC', 'LCM', 'VAM', 'MODI']:
                if method_name not in self.results:
                    continue
                result = self.results[method_name]
                solver = result['solver']
                sol = result['solution']
                cost = result['cost']
                is_best = (cost == best_cost)
                
                row += 2
                title_text = solver.method_used + " - Total Cost: $" + str(cost)
                if is_best:
                    title_text += " (OPTIMAL)"
                ws1.cell(row=row, column=1, value=title_text).font = Font(name='Segoe UI', bold=True, color='283593', size=11)
                
                # Solution matrix header
                row += 1
                ws1.cell(row=row, column=1, value="")
                for j in range(solver.num_demand):
                    ws1.cell(row=row, column=j+2, value="D" + str(j+1))
                ws1.cell(row=row, column=solver.num_demand+2, value="Supply")
                style_sub_header_row(ws1, row, solver.num_demand + 2)
                
                for i in range(solver.num_supply):
                    row += 1
                    ws1.cell(row=row, column=1, value="S" + str(i+1))
                    style_data_cell(ws1, row, 1, bold=True)
                    for j in range(solver.num_demand):
                        val = sol[i][j]
                        display = str(int(val)) if val == int(val) else "{:.1f}".format(val)
                        ws1.cell(row=row, column=j+2, value=display)
                        if val > 0:
                            style_data_cell(ws1, row, j+2, bold=True, fill=green_fill)
                        else:
                            style_data_cell(ws1, row, j+2)
                    ws1.cell(row=row, column=solver.num_demand+2, value=solver.supply[i])
                    style_data_cell(ws1, row, solver.num_demand+2, bold=True)
                
                # Demand row
                row += 1
                ws1.cell(row=row, column=1, value="Demand")
                style_data_cell(ws1, row, 1, bold=True)
                for j in range(solver.num_demand):
                    ws1.cell(row=row, column=j+2, value=solver.demand[j])
                    style_data_cell(ws1, row, j+2)
                ws1.cell(row=row, column=solver.num_demand+2, value="")
                style_data_cell(ws1, row, solver.num_demand+2)
                
                # Allocations
                row += 1
                ws1.cell(row=row, column=1, value="Allocations:").font = Font(name='Segoe UI', bold=True, size=10)
                summary = solver.get_solution_summary()
                for assign in summary['assignments']:
                    row += 1
                    a = assign['amount']
                    a_str = str(int(a)) if a == int(a) else "{:.1f}".format(a)
                    ws1.cell(row=row, column=1, value=assign['from'] + " -> " + assign['to'])
                    ws1.cell(row=row, column=2, value=a_str + " units")
                    ws1.cell(row=row, column=3, value="$" + str(assign['unit_cost']) + "/unit")
                    ws1.cell(row=row, column=4, value="$" + str(assign['total_cost']))
                    for c in range(1, 5):
                        style_data_cell(ws1, row, c)
            
            # 3. Comparison
            row += 3
            ws1.cell(row=row, column=1, value="3. Method Comparison").font = Font(name='Segoe UI', bold=True, color='1A237E', size=13)
            
            row += 1
            ws1.cell(row=row, column=1, value="Method")
            ws1.cell(row=row, column=2, value="Total Cost")
            ws1.cell(row=row, column=3, value="Allocations")
            ws1.cell(row=row, column=4, value="Status")
            style_header_row(ws1, row, 4)
            
            best_cost = min(self.results[k]['cost'] for k in self.results)
            for method_name in ['NWC', 'LCM', 'VAM', 'MODI']:
                if method_name not in self.results:
                    continue
                row += 1
                cost = self.results[method_name]['cost']
                is_best = (cost == best_cost)
                alloc_count = sum(1 for row2 in self.results[method_name]['solution'] for v in row2 if v > 0)
                
                ws1.cell(row=row, column=1, value=method_name)
                ws1.cell(row=row, column=2, value=cost)
                ws1.cell(row=row, column=3, value=alloc_count)
                ws1.cell(row=row, column=4, value="OPTIMAL" if is_best else "—")
                
                for c in range(1, 5):
                    style_data_cell(ws1, row, c, bold=is_best, fill=green_fill if is_best else None)
            
            # 4. Summary
            best_method = min(self.results, key=lambda k: self.results[k]['cost'])
            row += 3
            ws1.cell(row=row, column=1, value="4. Summary").font = Font(name='Segoe UI', bold=True, color='1A237E', size=13)
            
            row += 1
            ws1.cell(row=row, column=1, value="Optimal Method:").font = Font(name='Segoe UI', bold=True, size=10)
            ws1.cell(row=row, column=2, value=best_method).font = Font(name='Segoe UI', bold=True, color='2E7D32', size=10)
            
            row += 1
            ws1.cell(row=row, column=1, value="Minimum Total Cost:").font = Font(name='Segoe UI', bold=True, size=10)
            ws1.cell(row=row, column=2, value="$" + str(self.results[best_method]['cost'])).font = Font(name='Segoe UI', bold=True, color='2E7D32', size=10)
            
            row += 1
            ws1.cell(row=row, column=1, value="Methods Compared:").font = Font(name='Segoe UI', bold=True, size=10)
            ws1.cell(row=row, column=2, value=str(len(self.results))).font = normal_font
            
            # 5. Analysis
            row += 3
            ws1.cell(row=row, column=1, value="5. Analysis").font = Font(name='Segoe UI', bold=True, color='1A237E', size=13)
            
            # Back to top of analysis for styling
            ws1.merge_cells(start_row=row+1, start_column=1, end_row=row+20, end_column=4)
            analysis_cell = ws1.cell(row=row+1, column=1)
            
            costs_dict = {m: self.results[m]['cost'] for m in self.results}
            sorted_methods = sorted(costs_dict, key=lambda m: costs_dict[m])
            
            analysis_text = "Cost Efficiency Analysis:\n"
            for idx, m in enumerate(sorted_methods):
                c = costs_dict[m]
                if idx == 0:
                    analysis_text += "  - " + m + " achieved the optimal cost of $" + str(c) + ".\n"
                else:
                    pct = round((c - costs_dict[sorted_methods[0]]) / costs_dict[sorted_methods[0]] * 100, 1)
                    analysis_text += "  - " + m + ": $" + str(c) + " (" + str(pct) + "% higher).\n"
            
            first = sorted_methods[0]
            last = sorted_methods[-1]
            savings = costs_dict[last] - costs_dict[first]
            savings_pct = round((costs_dict[last] - costs_dict[first]) / costs_dict[last] * 100, 1)
            
            analysis_text += "\nKey Findings:\n"
            analysis_text += "  - Using " + first + " saves $" + str(savings) + " (" + str(savings_pct) + "%).\n"
            
            if 'MODI' in costs_dict and 'VAM' in costs_dict:
                if costs_dict['MODI'] == costs_dict['VAM']:
                    analysis_text += "  - MODI confirmed VAM was already optimal.\n"
                else:
                    analysis_text += "  - MODI improved VAM from $" + str(costs_dict['VAM']) + " to $" + str(costs_dict['MODI']) + ".\n"
            
            if total_supply != total_demand:
                analysis_text += "  - A dummy variable was added due to imbalance.\n"
            
            analysis_text += "\nRecommendation: Use " + first + " to minimize total cost to $" + str(costs_dict[first]) + "."
            
            analysis_cell.value = analysis_text
            analysis_cell.font = Font(name='Segoe UI', size=10)
            analysis_cell.alignment = Alignment(wrap_text=True, vertical='top')
            analysis_cell.fill = analysis_fill
            analysis_cell.border = thin_border
            
            # Footer
            row = row + 22
            ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws1.cell(row=row, column=1, value="Transportation Problem Solver - Prepared by Navid Naderpour,navid.naderpour@gmail.com").font = Font(name='Segoe UI', size=8, italic=True, color='757575')
            ws1.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            row += 1
            ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws1.cell(row=row, column=1, value="Generated on " + datetime.now().strftime("%Y-%m-%d at %H:%M:%S")).font = Font(name='Segoe UI', size=8, italic=True, color='757575')
            ws1.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # ===== Sheet 2: Raw Data =====
            ws2 = wb.create_sheet(title="Raw Data")
            
            ws2.column_dimensions['A'].width = 12
            ws2.column_dimensions['B'].width = 12
            ws2.column_dimensions['C'].width = 12
            ws2.column_dimensions['D'].width = 12
            ws2.column_dimensions['E'].width = 12
            
            # Supply
            ws2.cell(row=1, column=1, value="Supply").font = Font(name='Segoe UI', bold=True, size=11)
            for i, s in enumerate(self.entered_supply):
                ws2.cell(row=i+2, column=1, value=s)
            
            # Demand
            ws2.cell(row=1, column=3, value="Demand").font = Font(name='Segoe UI', bold=True, size=11)
            for j, d in enumerate(self.entered_demand):
                ws2.cell(row=j+2, column=3, value=d)
            
            # Cost matrix
            ws2.cell(row=1, column=5, value="Cost Matrix").font = Font(name='Segoe UI', bold=True, size=11)
            for i, cost_row in enumerate(self.entered_costs):
                for j, c in enumerate(cost_row):
                    ws2.cell(row=i+2, column=5+j, value=c)
            
            # Results for each method
            row2 = 1
            for method_name in ['NWC', 'LCM', 'VAM', 'MODI']:
                if method_name not in self.results:
                    continue
                result = self.results[method_name]
                solver = result['solver']
                sol = result['solution']
                cost = result['cost']
                
                row2 += 3
                ws2.cell(row=row2, column=1, value=method_name + " Solution").font = Font(name='Segoe UI', bold=True, size=10)
                row2 += 1
                
                for i in range(solver.num_supply):
                    for j in range(solver.num_demand):
                        if sol[i][j] > 0:
                            ws2.cell(row=row2, column=1, value="S" + str(i+1))
                            ws2.cell(row=row2, column=2, value="D" + str(j+1))
                            val = sol[i][j]
                            ws2.cell(row=row2, column=3, value=int(val) if val == int(val) else round(val, 1))
                            ws2.cell(row=row2, column=4, value=solver.costs[i][j])
                            ws2.cell(row=row2, column=5, value=solver.costs[i][j] * sol[i][j])
                            row2 += 1
                
                ws2.cell(row=row2, column=1, value="Total Cost").font = Font(name='Segoe UI', bold=True, size=10)
                ws2.cell(row=row2, column=2, value=cost).font = Font(name='Segoe UI', bold=True, color='2E7D32', size=10)
                row2 += 1
            
            # Save
            from tkinter import filedialog
            default_name = "transport_report_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
            filename = filedialog.asksaveasfilename(
                initialdir=desktop,
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Excel Report"
            )
            if not filename:
                return  # User cancelled
            wb.save(filename)
            
            messagebox.showinfo("Excel Generated", "Excel report saved to:\n" + filename)
            
        except Exception as e:
            messagebox.showerror("Excel Error", "Could not generate Excel:\n" + str(e) + 
                               "\n\nMake sure 'openpyxl' is installed:\npip install openpyxl")

# ============================================================
# Main Entry Point
# ============================================================

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1 and sys.argv[1] == "test":
        # Console test mode
        supply = [10, 15, 20]
        demand = [8, 12, 10, 15]
        costs = [
            [3, 7, 2, 5],
            [4, 1, 8, 6],
            [9, 3, 4, 2]
        ]
        
        print("=" * 60)
        print("TRANSPORTATION PROBLEM - ALL 4 METHODS")
        print("=" * 60)
        print("Origins (Supply): " + str(supply))
        print("Destinations (Demand): " + str(demand))
        print("Cost Matrix:")
        for i, row in enumerate(costs):
            print("  S" + str(i+1) + ": " + str(row))
        
        s1 = TransportSolver(supply, demand, costs)
        sol1, cost1, _ = s1.nwc()
        
        s2 = TransportSolver(supply, demand, costs)
        sol2, cost2, _ = s2.least_cost()
        
        s3 = TransportSolver(supply, demand, costs)
        sol3, cost3, _ = s3.vogel()
        
        s4 = TransportSolver(supply, demand, costs)
        sol4, cost4, _ = s4.modi('vogel')
        
        print("\n" + "=" * 60)
        print("FINAL COMPARISON - ALL METHODS")
        print("=" * 60)
        print("Method                              Total Cost")
        print("-" * 50)
        print("1. NWC (North-West Corner)          " + str(cost1))
        print("2. LCM (Least Cost)                 " + str(cost2))
        print("3. VAM (Vogel)                      " + str(cost3))
        print("4. MODI (from VAM) - OPTIMAL        " + str(cost4))
        print("-" * 50)
        costs_dict = {
            "NWC": cost1,
            "LCM": cost2,
            "VAM": cost3,
            "MODI": cost4
        }

        best_method = min(costs_dict, key=costs_dict.get)
        best_cost = costs_dict[best_method]

        print("\nBEST = " + best_method + " with Total Cost = " + str(best_cost))
        
        print("\nVAM SOLUTION:")
        s3.print_solution()
        print("\nMODI SOLUTION:")
        s4.print_solution()
        
    else:
        # Launch GUI
        root = tk.Tk()
        app = TransportGUI(root)
        root.mainloop()
           





